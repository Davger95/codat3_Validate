from openpyxl import load_workbook
from pathlib import Path

path = Path('/home/Dave/.openclaw/workspace-datadict/P_workspace_codat3_validator/templates/test_files/Strukturvorlage_DataDictionary_test1.xlsx')
wb = load_workbook(path)

# Header should remain authored and valid
ws = wb['Header']
ws.cell(4,2).value = 'bdch'
ws.cell(5,2).value = 'datenkatalog_flaechenmanagement'
ws.cell(6,2).value = 'Datenkatalog Flächenmanagement'
ws.cell(8,2).value = 'Data Dictionary Area Management'
ws.cell(9,2).value = '1.0.0'
ws.cell(10,2).value = 'https://example.com/bdch/datenkatalog_flaechenmanagement'
ws.cell(13,2).value = 'Active'

# Classes
ws = wb['Classes']
ws.cell(1,31).value = 'RelatedDocumentName (EN)'
if ws.cell(8,31).value == 'organisation':
    ws.cell(8,31).value = 'Organisation'

# Properties
ws = wb['Properties']
ws.cell(1,17).value = 'EnumerationDesignation (EN)'
# use designation instead of old enum id linkage for the authored sample
ws.cell(8,17).value = 'Raumbezeichnung SIA 380/2015'

# Values
ws = wb['Values']
ws.cell(8,9).value = 'Raumbezeichnung SIA 380/2015'

# Documents
ws = wb['Documents']
ws.cell(1,4).value = 'GUID/URI'
ws.cell(1,11).value = 'Security level/Sicherheitsstufe/Niveau de sécurité/Livello di sicurezza'
ws.cell(1,12).value = 'Accessibility/Zugänglichkeit/Accessibilité/Accessibilità'
ws.delete_cols(15, 1)
ws.cell(1,15).value = 'Governance'
ws.cell(1,16).value = 'Status'
ws.cell(1,17).value = 'Version date'
ws.cell(1,18).value = 'Provenance (PROV)'
# restore row 8 content after O-column removal shift
ws.cell(8,4).value = None
ws.cell(8,5).value = 'organisation'
ws.cell(8,6).value = 'ORG'
ws.cell(8,7).value = 'Organisation'
ws.cell(8,8).value = 'Organisation'
ws.cell(8,11).value = 'intern'
ws.cell(8,12).value = 'Zugänglich unentgeltlich'
ws.cell(8,13).value = '1.0'
ws.cell(8,14).value = 'bdch'
ws.cell(8,16).value = 'Active'
ws.cell(8,17).value = '2026-06-18T15:30+02:00'
ws.cell(8,18).value = 'bdch'

wb.save(path)
print('authored sample synced')
