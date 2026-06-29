from __future__ import annotations

from dataclasses import replace
from pathlib import Path
from typing import Iterable
from urllib.parse import quote
import re
import openpyxl

from .excel_reader import (
    DataDictionary,
    DictionaryMeta,
    DDAllowedValue,
    DDClass,
    DDClassProperty,
    DDConceptRelation,
    DDProperty,
)


def slugify(value: str) -> str:
    s = (value or "").strip().lower()
    repl = {
        "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss",
        "à": "a", "á": "a", "â": "a", "ã": "a", "å": "a",
        "è": "e", "é": "e", "ê": "e", "ë": "e",
        "ì": "i", "í": "i", "î": "i", "ï": "i",
        "ò": "o", "ó": "o", "ô": "o", "õ": "o",
        "ù": "u", "ú": "u", "û": "u",
        "ç": "c", "ñ": "n",
        "_": "-",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r"[^a-z0-9-]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def _str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _split_values(raw: str | None) -> list[str]:
    if not raw:
        return []
    txt = str(raw).strip()
    if not txt:
        return []
    parts = [p.strip().strip('"') for p in re.split(r"[;,]", txt)]
    return [p for p in parts if p]


def _safe_uri(base: str, code: str) -> str:
    return f"{base}{quote(code, safe='')}"


def _normalize_ifc_entity(entity: str | None, predefined: str | None) -> str | None:
    entity = _str(entity)
    predefined = _str(predefined)
    if not entity:
        return None
    if predefined and entity.endswith(predefined):
        candidate = entity[: -len(predefined)]
        if candidate.startswith("Ifc") and len(candidate) > 3:
            return candidate
    return entity


def _detect_header_row(ws, required_marker: str) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True), start=1):
        vals = [_str(v) or "" for v in row]
        if required_marker in vals:
            return i
    raise ValueError(f"Could not find header row containing {required_marker!r} in sheet {ws.title}")


def _row_has_meaningful_content(row) -> bool:
    meaningful = []
    for v in row:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        meaningful.append(s)
    if not meaningful:
        return False
    joined = ' '.join(meaningful).lower()
    guidance_markers = [
        'should reference',
        'should match',
        'fill if',
        'leave empty',
        'required human-readable',
        'required registered source code',
        'reference/list-based validation',
        'optional external',
        'optional proprietary/helper field',
        'optional governance notes',
        'optional maintenance notes',
        'system generated canonical',
        'required canonical',
        'validator should',
    ]
    if any(marker in joined for marker in guidance_markers):
        return False
    return True


def _row_dicts(ws, header_row: int, data_start: int) -> list[dict[str, object]]:
    headers = [_str(c.value) or "" for c in ws[header_row]]
    out = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if _row_has_meaningful_content(row):
            out.append(dict(zip(headers, row)))
    return out


def _default_meta(path: Path) -> DictionaryMeta:
    stem = path.stem.lower()
    if "kbob" in stem:
        code = "kbob"
        uri = "https://www.kbob.admin.ch/dictionary/kbob"
        de = "KBOB Datenkatalog"
        en = "KBOB Data Dictionary"
        dictionary_code = "kbob"
    elif "bdch" in stem:
        code = "bdch"
        uri = "https://bauen-digital.ch/dictionary/bdch"
        de = "bDCH Datenkatalog"
        en = "bDCH Data Dictionary"
        dictionary_code = "bdch"
    else:
        code = "he-sem"
        dictionary_code = slugify(path.stem) or "he-dd"
        uri = f"https://he-sem.ch/dictionary/{dictionary_code}"
        de = f"HE-DD Datenkatalog {path.stem}"
        en = f"HE-DD Data Dictionary {path.stem}"
    return DictionaryMeta(
        org_code=code,
        org_name_de=code.upper(),
        org_name_en=code.upper(),
        dd_uri=uri,
        dd_version="0.1",
        dd_status="Preview",
        raw={
            "OrganizationCode": code,
            "DictionaryCode": dictionary_code,
            "DictionaryUri": uri,
            "DictionaryName (DE)": de,
            "DictionaryName (EN)": en,
            "DictionaryVersion": "0.1",
            "Version": "0.1",
            "Status": "Preview",
        },
    )


def _parse_dictionary_if_present(wb, path: Path) -> DictionaryMeta:
    if "Dictionary_core" in wb.sheetnames:
        ws = wb["Dictionary_core"]
    elif "Dictionary" in wb.sheetnames:
        ws = wb["Dictionary"]
    else:
        return _default_meta(path)
    meta = _default_meta(path)
    for row in ws.iter_rows(min_row=2, values_only=True):
        field_name = _str(row[0]) if len(row) > 0 else None
        value = _str(row[1]) if len(row) > 1 else None
        if not field_name or value is None:
            continue
        meta.raw[field_name] = value
        fl = field_name.lower()
        if fl == "organizationcode":
            meta.org_code = value
        elif fl in ("dictionaryuri",):
            meta.dd_uri = value
        elif fl in ("version", "dictionaryversion"):
            meta.dd_version = value
        elif fl == "status":
            meta.dd_status = value
        elif fl == "countriesofuse":
            meta.countries = value
        elif field_name == "OrganizationName_DE":
            meta.org_name_de = value
        elif field_name == "OrganizationName_EN":
            meta.org_name_en = value
    return meta


def _lindas_base(meta: DictionaryMeta) -> str:
    org1 = (meta.raw.get("OrganizationCodeLindas") or meta.raw.get("OrganizationCode") or meta.org_code or "org").upper()
    org2 = (meta.raw.get("OrganizationSubCode") or meta.raw.get("OrganizationCode") or meta.org_code or org1).upper()
    dd_code = (meta.raw.get("DictionaryCode") or meta.raw.get("DictionaryName (EN)") or meta.raw.get("DictionaryName (DE)") or "dd").strip()
    dd_code = slugify(dd_code).upper().replace('-', '_')
    return f"https://lindas.admin.ch/{org1}/{org2}/{dd_code}/"


def _parse_classes(ws, meta: DictionaryMeta) -> list[DDClass]:
    rows = _row_dicts(ws, header_row=3, data_start=10)
    base = f"{_lindas_base(meta)}class/"
    classes = []
    for r in rows:
        designation = _str(r.get("Designation")) or _str(r.get("Bezeichnung")) or ""
        source_objekt_id = _str(r.get("Objekt-ID"))
        bezeichnung = _str(r.get("Bezeichnung")) or ""
        beschreibung = _str(r.get("Beschreibung")) or _str(r.get("Description")) or ""
        if not any([source_objekt_id, designation, bezeichnung, beschreibung]):
            continue
        code = source_objekt_id or slugify(designation)
        if not code:
            continue
        owned_uri = _str(r.get("GUID/URI")) or _safe_uri(base, code)
        predefined = _str(r.get("PredefinedType"))
        entity_raw = _str(r.get("IfcObject Entity"))
        cls = DDClass(
            code=code,
            class_type="Class",
            name_de=_str(r.get("Bezeichnung")) or designation,
            name_fr="",
            name_en=designation,
            definition_de=_str(r.get("Beschreibung")) or "",
            definition_fr=_str(r.get("Description")) or "",
            owned_uri=owned_uri,
            parent_class_code=_str(r.get("Objekt-Einordnung")),
            ifc_entity_code=_normalize_ifc_entity(entity_raw, predefined),
            ifc_predefined_type=predefined,
            ifc_uri=_str(r.get("IFC URI")),
            rds_reference=_str(r.get("Identification")),
            crb_code=_str(r.get("Name")),
            status="Preview",
            document_reference=_str(r.get("Source")),
            countries_of_use="CH",
        )
        setattr(cls, "ifc_type_object_entity_code", _str(r.get("IfcTypeObject Entity")))
        setattr(cls, "generated_objekt_id", slugify(designation) if designation else None)
        classes.append(cls)
    return classes


def _reconcile_class_codes_from_matrix(ws, classes: list[DDClass]) -> list[DDClass]:
    object_labels = [(_str(v) or "") for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    object_ids = [(_str(v) or "") for v in next(ws.iter_rows(min_row=5, max_row=5, values_only=True))]
    by_de = {c.name_de: c for c in classes if c.name_de}
    by_en = {c.name_en: c for c in classes if c.name_en}
    out = []
    seen = set()
    for idx, obj_id in enumerate(object_ids, start=1):
        if idx < 8 or not obj_id or obj_id == 'Objekt-ID':
            continue
        label = object_labels[idx - 1] if idx - 1 < len(object_labels) else ''
        cls = by_de.get(label) or by_en.get(label)
        if cls and cls.code not in seen:
            cls.code = obj_id
            seen.add(cls.code)
    return classes


def _infer_property_value_kind(values: list[str]) -> str:
    return "List" if values else "Single"


def _parse_properties(ws, meta: DictionaryMeta) -> tuple[list[DDProperty], list[DDAllowedValue]]:
    rows = _row_dicts(ws, header_row=3, data_start=4)
    lindas_base = _lindas_base(meta)
    base_prop = f"{lindas_base}property/"
    base_val = f"{lindas_base}allowed-value/"
    props = []
    avs = []
    for r in rows:
        name_en = _str(r.get("Property")) or _str(r.get("Merkmal")) or ""
        code = _str(r.get("Merkmal-ID")) or slugify(name_en)
        vals = _split_values(_str(r.get("Werteliste")))
        prop = DDProperty(
            code=code,
            name_de=_str(r.get("Merkmal")) or "",
            name_fr="",
            name_en=name_en,
            definition_de=_str(r.get("Beschreibung")) or "",
            definition_fr="",
            owned_uri=_str(r.get("GUID/URI")) or _safe_uri(base_prop, code),
            data_type=_str(r.get("DataType\n(Base Type)")) or _str(r.get("DataType (Base Type)")) or "STRING",
            data_type_ifc=_str(r.get("DataType\n(IFC)")) or _str(r.get("DataType (IFC)")),
            property_value_kind=_infer_property_value_kind(vals),
            unit_label=_str(r.get("Einheiten")),
            enumeration_values=",".join(vals) if vals else None,
            ifc_property_uri=None,
            ifc_pset_uri=_str(r.get("IfcPropertySet (Pset)")),
            property_set_name=_str(r.get("Benutzerdefiniertes PropertySet")) or _str(r.get("Merkmalsgruppe\n(group of property)")) or _str(r.get("Merkmalsgruppe (group of property)")),
            rds_reference=_str(r.get("Zuordnung")),
            status="Preview",
        )
        props.append(prop)
        for idx, v in enumerate(vals, start=1):
            val_code = slugify(v)
            avs.append(DDAllowedValue(
                property_code=code,
                code=val_code,
                value_de=v,
                value_fr="",
                value_en=v,
                definition_de=None,
                owned_uri=_safe_uri(base_val, f"{code}/{val_code}"),
                sort_number=idx,
                status="Preview",
            ))
    return props, avs


def _parse_documents(ws, meta: DictionaryMeta) -> list[dict]:
    rows = _row_dicts(ws, header_row=3, data_start=9)
    docs = []
    for r in rows:
        source_code = _str(r.get("SourceCode")) or _str(r.get("Zuordnung")) or ""
        document_code = _str(r.get("DocumentCode")) or _str(r.get("Identification")) or ""
        document_label = _str(r.get("DocumentLabel")) or _str(r.get("Name")) or ""
        document_group_code = _str(r.get("DocumentGroupCode")) or _str(r.get("Identification_1")) or ""
        document_group_label = _str(r.get("DocumentGroupLabel")) or _str(r.get("Dokumentgruppe")) or ""
        dokument_id = _str(r.get("Dokument-ID")) or slugify(f"{source_code}-{document_code}")
        item = dict(r)
        item["SourceCode"] = source_code
        item["Dokument-ID"] = dokument_id
        item["DocumentCode"] = document_code
        item["DocumentLabel"] = document_label
        item["DocumentGroupCode"] = document_group_code
        item["DocumentGroupName"] = document_group_label
        base_doc = f"{_lindas_base(meta)}document/"
        item["OwnedUri"] = _safe_uri(base_doc, dokument_id)
        docs.append(item)
    return docs


def _parse_objects_v20260619(ws, meta: DictionaryMeta) -> list[DDClass]:
    rows = _row_dicts(ws, header_row=3, data_start=10)
    base = f"{_lindas_base(meta)}class/"
    classes = []
    for r in rows:
        if not any([
            _str(r.get("Objekt-ID")),
            _str(r.get("Designation")),
            _str(r.get("Bezeichnung")),
            _str(r.get("Beschreibung")),
        ]):
            continue
        code = _str(r.get("Objekt-ID")) or slugify(_str(r.get("Designation")) or _str(r.get("Bezeichnung")) or "")
        if not code:
            continue
        cls = DDClass(
            code=code,
            class_type="Class",
            name_de=_str(r.get("Bezeichnung")) or "",
            name_fr="",
            name_en=_str(r.get("Designation")) or _str(r.get("Bezeichnung")) or "",
            definition_de=_str(r.get("Beschreibung")) or "",
            definition_fr="",
            owned_uri=_str(r.get("GUID/URI")) or _safe_uri(base, code),
            parent_class_code=_str(r.get("Objekt-Einordnung ")) or _str(r.get("Objekt-Einordnung")),
            ifc_entity_code=_str(r.get("IfcObject Entity")),
            ifc_predefined_type=None,
            ifc_uri=_str(r.get("Objekte.IFC_URI")) or _str(r.get("IFC_URI")) or _str(r.get("GUID/URI_1")) or _str(r.get("IFC URI")),
            rds_reference=None,
            crb_code=None,
            status="Preview",
            document_reference=_str(r.get("Hinweise zur Handhabung")),
            countries_of_use="CH",
        )
        setattr(cls, "ifc_type_object_entity_code", _str(r.get("IfcTypeObject Entity")))
        classes.append(cls)
    return classes


def _parse_properties_v20260619(ws, value_ws, meta: DictionaryMeta) -> tuple[list[DDProperty], list[DDAllowedValue]]:
    rows = _row_dicts(ws, header_row=2, data_start=8)
    value_rows = _row_dicts(value_ws, header_row=3, data_start=10) if value_ws is not None else []
    value_map = {}
    for vr in value_rows:
        raw_id = _str(vr.get("Werte-ID"))
        if raw_id:
            value_map[raw_id] = vr
    value_base = f"{_lindas_base(meta)}allowed-value/"
    prop_base = f"{_lindas_base(meta)}property/"
    props = []
    avs = []
    for r in rows:
        code = _str(r.get("Merkmal-ID"))
        if not code:
            continue
        value_list_id = _str(r.get("Werteliste-ID"))
        value_row = value_map.get(value_list_id or "")
        enum_values_raw = _str(value_row.get("EnumerationValues\n(EN)")) if value_row else None
        vals = _split_values(enum_values_raw)
        prop = DDProperty(
            code=code,
            name_de=_str(r.get("Merkmal")) or "",
            name_fr="",
            name_en=_str(r.get("Property")) or _str(r.get("Merkmal")) or "",
            definition_de=_str(r.get("Beschreibung")) or "",
            definition_fr="",
            owned_uri=_str(r.get("GUID/URI")) or _safe_uri(prop_base, code),
            data_type=_str(r.get("DataType\n(Base Type)")) or "STRING",
            data_type_ifc=_str(r.get("DataType\n(IFC)")),
            property_value_kind=_infer_property_value_kind(vals),
            unit_label=_str(value_row.get("Einheiten")) if value_row else None,
            enumeration_values=",".join(vals) if vals else None,
            ifc_property_uri=_str(r.get("GUID/URI_1")),
            ifc_pset_uri=_str(r.get("IfcPropertySet (Pset)")),
            property_set_name=_str(r.get("Benutzerdefiniertes PropertySet")) or _str(r.get("Governance (Merkmale)")),
            rds_reference=_str(r.get("Hinweise zur Handhabung")),
            status="Preview",
        )
        setattr(prop, "value_list_id", value_list_id)
        setattr(prop, "raw_ifc_qto", _str(r.get("IfcQuantitySet (Qto)")))
        setattr(prop, "merkmalsgruppe", _str(r.get("Benutzerdefiniertes PropertySet")))
        props.append(prop)
        if value_row and vals:
            for idx, v in enumerate(vals, start=1):
                val_code = slugify(v)
                avs.append(DDAllowedValue(
                    property_code=code,
                    code=val_code,
                    value_de=v,
                    value_fr=v,
                    value_en=v,
                    definition_de=_str(value_row.get("Bezeichnung")),
                    owned_uri=_safe_uri(value_base, f"{code}/{val_code}"),
                    sort_number=idx,
                    status=_str(value_row.get("Status")) or "Preview",
                ))
    return props, avs


def _parse_documents_v20260619(ws, meta: DictionaryMeta) -> list[dict]:
    docs = []
    for row_idx in range(10, ws.max_row + 1):
        source_code = _str(ws.cell(row_idx, 2).value) or ""
        if not any(_str(ws.cell(row_idx, c).value) for c in range(1, min(ws.max_column, 17) + 1)):
            continue
        classif_identification = _str(ws.cell(row_idx, 4).value)
        classif_name = _str(ws.cell(row_idx, 5).value)
        revision = _str(ws.cell(row_idx, 6).value)
        owner = _str(ws.cell(row_idx, 7).value)
        location = _str(ws.cell(row_idx, 8).value)
        document_code = _str(ws.cell(row_idx, 9).value)
        document_label = _str(ws.cell(row_idx, 10).value)
        group_code = _str(ws.cell(row_idx, 12).value)
        group_label = _str(ws.cell(row_idx, 13).value)
        dokument_id = slugify(f"{source_code}-{document_code or classif_identification or classif_name or row_idx}")
        item = {
            "SourceCode": source_code,
            "Dokument-ID": dokument_id,
            "DocumentCode": document_code,
            "DocumentLabel": document_label,
            "DocumentGroupCode": group_code,
            "DocumentGroupName": group_label,
            "OwnedUri": _safe_uri(f"{_lindas_base(meta)}document/", dokument_id),
            "Document Identification": classif_identification,
            "Dokument Name": classif_name,
            "Revision": revision,
            "Owner": owner,
            "Location": location,
            "Dokument URI": classif_identification if (classif_identification and str(classif_identification).startswith('http')) else None,
        }
        docs.append(item)
    return docs


def _parse_matrix_v20260619(ws, properties: list[DDProperty]) -> list[DDClassProperty]:
    prop_ids = []
    for col_idx in range(5, ws.max_column + 1):
        prop_code = _str(ws.cell(3, col_idx).value)
        if prop_code:
            prop_ids.append((col_idx, prop_code))
    cps = []
    prop_index = {p.code: p for p in properties}
    for row_idx in range(8, ws.max_row + 1):
        class_code = _str(ws.cell(row_idx, 1).value)
        if not class_code:
            continue
        for col_idx, prop_code in prop_ids:
            raw = _str(ws.cell(row_idx, col_idx).value)
            if not raw:
                continue
            marker = raw.strip()
            lower = marker.lower()
            override = None if lower == 'x' else marker
            p = prop_index.get(prop_code)
            cps.append(DDClassProperty(
                class_code=class_code,
                property_code=prop_code,
                property_set_name=(p.property_set_name if p else None),
                is_required=True,
                is_writable=True,
                allowed_values_override=override,
            ))
    return cps


def load_he_dd_v20260619(path: Path) -> DataDictionary:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    meta = _default_meta(path)
    meta.raw.update({
        "DictionaryCode": "DD_FM",
        "OrganizationCode": meta.org_code or "bdch",
    })
    classes = _parse_objects_v20260619(wb["Objekte"], meta)
    properties, allowed_values = _parse_properties_v20260619(wb["Merkmale_Merkmalsgruppen"], wb["Werte"], meta)
    class_properties = _parse_matrix_v20260619(wb["Data Template AreaMgmt"], properties)
    dd = DataDictionary(
        source_file=path,
        meta=meta,
        classes=classes,
        properties=properties,
        class_properties=class_properties,
        allowed_values=allowed_values,
        concept_relations=[],
    )
    setattr(dd, "documents", _parse_documents_v20260619(wb["Dokumente_Dokumentgruppen"], meta))
    return dd


def _parse_objects_abgeglichen(ws, meta: DictionaryMeta) -> list[DDClass]:
    rows = _row_dicts(ws, header_row=3, data_start=10)
    base = f"{_lindas_base(meta)}class/"
    classes = []
    for r in rows:
        code = _str(r.get("Objekt-ID"))
        if not code:
            continue
        cls = DDClass(
            code=code,
            class_type="Class",
            name_de=_str(r.get("Bezeichnung (DE)")) or _str(r.get("Bezeichnung")) or "",
            name_fr=_str(r.get("Désignation (FR)")) or _str(r.get("FR")) or "",
            name_en=_str(r.get("Designation (EN)")) or _str(r.get("Designation")) or _str(r.get("Bezeichnung (DE)")) or "",
            definition_de=_str(r.get("Beschreibung (DE)")) or _str(r.get("Beschreibung")) or "",
            definition_fr=_str(r.get("Description (FR)")) or _str(r.get("Beschreibung (FR)")) or "",
            owned_uri=_str(r.get("GUID/URI")) or _safe_uri(base, code),
            parent_class_code=_str(r.get("Objekt-Einordnung ")) or _str(r.get("Objekt-Einordnung")),
            ifc_entity_code=_normalize_ifc_entity(_str(r.get("IfcObject Entity")), _str(r.get("PredefinedType"))),
            ifc_predefined_type=_str(r.get("PredefinedType")),
            ifc_uri=_str(r.get("Objekte.IFC_URI")) or _str(r.get("IFC_URI")) or _str(r.get("GUID/URI_1")) or _str(r.get("IFC URI")),
            rds_reference=_str(r.get("Klassifikationen")),
            crb_code=None,
            status=_str(r.get("Status")) or "Preview",
            document_reference=_str(r.get("Herkunft (PROV)")),
            countries_of_use="CH",
        )
        setattr(cls, "name_it", _str(r.get("Designazione (IT)")) or _str(r.get("IT")) or "")
        setattr(cls, "definition_it", _str(r.get("Descrizione (IT)")) or _str(r.get("Beschreibung (IT)")) or "")
        setattr(cls, "ifc_type_object_entity_code", _str(r.get("IfcTypeObject Entity")))
        setattr(cls, "object_type", _str(r.get("ObjectType")))
        setattr(cls, "status", _str(r.get("Status")) or "Preview")
        setattr(cls, "version_date", _str(r.get("Versionsdatum")))
        setattr(cls, "related_document", _str(r.get("Objekte.RelatedDocument (Document-ID)")) or _str(r.get("RelatedDocument (Document-ID)")) or _str(r.get("RelatedDocument")) or 'Organisation')
        classes.append(cls)
    return classes


def _parse_properties_abgeglichen(ws, value_ws, meta: DictionaryMeta) -> tuple[list[DDProperty], list[DDAllowedValue]]:
    rows = _row_dicts(ws, header_row=2, data_start=8)
    value_rows = _row_dicts(value_ws, header_row=3, data_start=10) if value_ws is not None else []
    value_map = {}
    for vr in value_rows:
        raw_id = _str(vr.get("Werteliste-ID")) or _str(vr.get("Werte-ID"))
        if raw_id:
            value_map[raw_id] = vr
    value_base = f"{_lindas_base(meta)}allowed-value/"
    prop_base = f"{_lindas_base(meta)}property/"
    props = []
    avs = []
    for r in rows:
        code = _str(r.get("Merkmal-ID")) or _str(r.get("Merkmal-Code"))
        if not code:
            continue
        value_list_id = _str(r.get("Werteliste-ID"))
        value_row = value_map.get(value_list_id or "")
        enum_values_raw = _str(value_row.get("EnumerationValues\n(EN)")) if value_row else None
        if not enum_values_raw and value_row:
            enum_values_raw = _str(value_row.get("Werteliste\n(DE)")) or _str(value_row.get("Lista valori\n(IT)")) or _str(value_row.get("Liste de valeurs\n(FR)"))
        vals = _split_values(enum_values_raw)
        prop = DDProperty(
            code=code,
            name_de=_str(r.get("Bezeichnung (DE)")) or _str(r.get("Merkmal")) or "",
            name_fr=_str(r.get("Description (FR)")) or _str(r.get("FR")) or _str(r.get("Désignation (FR)")) or "",
            name_en=_str(r.get("Designation (EN)")) or _str(r.get("Property")) or _str(r.get("Bezeichnung (DE)")) or _str(r.get("Merkmal")) or "",
            definition_de=_str(r.get("Beschreibung (DE)")) or _str(r.get("Beschreibung")) or "",
            definition_fr=_str(r.get("Description (FR)")) or _str(r.get("Beschreibung (FR)")) or "",
            owned_uri=_str(r.get("GUID/URI")) or _safe_uri(prop_base, code),
            data_type=_str(r.get("DataType\n(Base Type)")) or "STRING",
            data_type_ifc=_str(r.get("DataType\n(IFC)")),
            property_value_kind=_infer_property_value_kind(vals),
            unit_label=_str(value_row.get("Einheiten")) if value_row else None,
            enumeration_values=",".join(vals) if vals else None,
            ifc_property_uri=_str(r.get("GUID/URI_1")),
            ifc_pset_uri=_str(r.get("IfcPropertySet (Pset)\nIfcQuantitySet (Qto)")),
            property_set_name=None,
            rds_reference=None,
            status=_str(r.get("Status")) or "Preview",
        )
        setattr(prop, "name_it", _str(r.get("Descrizione (IT)")) or _str(r.get("IT")) or _str(r.get("Designazione (IT)")) or "")
        setattr(prop, "definition_it", _str(r.get("Descrizione (IT)")) or _str(r.get("Beschreibung (IT)")) or "")
        setattr(prop, "value_list_id", value_list_id)
        setattr(prop, "raw_ifc_qto", _str(r.get("IfcPropertySet (Pset)\nIfcQuantitySet (Qto)")))
        setattr(prop, "merkmalsgruppe", None)
        props.append(prop)
        if value_row and vals:
            value_de_list = _split_values(_str(value_row.get("Werteliste\n(DE)")))
            value_fr_list = _split_values(_str(value_row.get("Liste de valeurs\n(FR)")))
            value_it_list = _split_values(_str(value_row.get("Lista valori\n(IT)")))
            for idx, v in enumerate(vals, start=1):
                val_code = slugify(v)
                av = DDAllowedValue(
                    property_code=code,
                    code=val_code,
                    value_de=value_de_list[idx - 1] if idx - 1 < len(value_de_list) else v,
                    value_fr=value_fr_list[idx - 1] if idx - 1 < len(value_fr_list) else v,
                    value_en=v,
                    definition_de=_str(value_row.get("Bezeichnung")),
                    owned_uri=_safe_uri(value_base, f"{code}/{val_code}"),
                    sort_number=idx,
                    status=_str(value_row.get("Status")) or "Preview",
                )
                setattr(av, "value_it", value_it_list[idx - 1] if idx - 1 < len(value_it_list) else v)
                avs.append(av)
    return props, avs


def _parse_documents_abgeglichen(ws, meta: DictionaryMeta) -> list[dict]:
    docs = []
    header_row = None
    for marker in ('DocumentName (EN)', 'DocumentName', 'DocumentationName', 'Bezeichnung (DE)'):
        try:
            header_row = _detect_header_row(ws, marker)
            break
        except ValueError:
            continue
    if header_row is None:
        raise ValueError(f"Could not find document header row in sheet {ws.title}")
    rows = _row_dicts(ws, header_row=header_row, data_start=header_row + 7)
    for row_idx, r in enumerate(rows, start=header_row + 1):
        document_uri = _str(r.get("URI"))
        document_identification = _str(r.get("Document-ID")) or _str(r.get("Document-ID\nIdentification"))
        document_code = _str(r.get("Document-Code")) or _str(r.get("Document-Code\nIdentification"))
        document_name = _str(r.get("DocumentName (EN)")) or _str(r.get("DocumentName")) or _str(r.get("DocumentationName"))
        security_level = _str(r.get("Sicherheitsstufe"))
        accessibility = _str(r.get("Zugänglichkeit"))
        revision = _str(r.get("Revision"))
        owner = _str(r.get("\tDocumentOwner")) or _str(r.get("DocumentOwner"))
        if not any([document_uri, document_identification, document_code, document_name, security_level, accessibility, revision, owner]):
            continue
        dokument_id = slugify(f"{document_identification or document_code or document_name or row_idx}")
        item = {
            "SourceCode": owner or "Organisation",
            "Dokument-ID": dokument_id,
            "DocumentCode": document_code,
            "DocumentLabel": document_name,
            "DocumentGroupCode": security_level,
            "DocumentGroupName": accessibility,
            "OwnedUri": _safe_uri(f"{_lindas_base(meta)}document/", dokument_id),
            "Document Identification": document_identification,
            "Dokument Name": document_name,
            "DocumentationName": document_name,
            "DocumentName": document_name,
            "DocumentName (DE)": _str(r.get("Bezeichnung (DE)")) or _str(r.get("DocumentName (DE)")) or _str(r.get("DE")),
            "DocumentationName (FR)": _str(r.get("Désignation (FR)")) or _str(r.get("DocumentName (FR)")) or _str(r.get("DocumentationName (FR)")) or _str(r.get("FR")),
            "DocumentationName (IT)": _str(r.get("Designazione (IT)")) or _str(r.get("DocumentName (IT)")) or _str(r.get("DocumentationName (IT)")) or _str(r.get("IT")),
            "Revision": revision,
            "Owner": owner,
            "Location": None,
            "Dokument URI": document_uri,
            "Sicherheitsstufe": security_level,
            "Zugänglichkeit": accessibility,
        }
        docs.append(item)
    return docs


def _parse_merkmalsgruppen_abgeglichen(ws) -> set[str]:
    groups = set()
    for row_idx in range(8, ws.max_row + 1):
        for col_idx in (4, 5, 6, 7):
            value = _str(ws.cell(row_idx, col_idx).value)
            if value:
                groups.add(value)
    return groups


def _parse_matrix_abgeglichen(ws, properties: list[DDProperty], allowed_group_refs: set[str]) -> list[DDClassProperty]:
    prop_ids = []
    for col_idx in range(9, ws.max_column + 1):
        prop_code = _str(ws.cell(3, col_idx).value)
        if prop_code:
            prop_ids.append((col_idx, prop_code))
    cps = []
    prop_index = {p.code: p for p in properties}
    for row_idx in range(8, ws.max_row + 1):
        class_code = _str(ws.cell(row_idx, 1).value)
        if not class_code:
            continue
        property_group = _str(ws.cell(row_idx, 3).value)
        for col_idx, prop_code in prop_ids:
            raw = _str(ws.cell(row_idx, col_idx).value)
            if not raw:
                continue
            marker = raw.strip()
            lower = marker.lower()
            override = None if lower == 'x' else marker
            p = prop_index.get(prop_code)
            cps.append(DDClassProperty(
                class_code=class_code,
                property_code=prop_code,
                property_set_name=property_group if property_group in allowed_group_refs else (p.property_set_name if p else property_group),
                is_required=True,
                is_writable=True,
                allowed_values_override=override,
            ))
    return cps


def load_he_dd_abgeglichen(path: Path) -> DataDictionary:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    meta = _default_meta(path)
    meta.raw.update({
        "DictionaryCode": slugify(path.stem).upper().replace('-', '_'),
        "OrganizationCode": meta.org_code or "bdch",
    })
    classes = _parse_objects_abgeglichen(wb["Objekte"], meta)
    properties, allowed_values = _parse_properties_abgeglichen(wb["Merkmale"], wb["Werte"], meta)
    group_sheet = "Merkmalgruppen" if "Merkmalgruppen" in wb.sheetnames else None
    group_refs = _parse_merkmalsgruppen_abgeglichen(wb[group_sheet]) if group_sheet else set()
    class_properties = _parse_matrix_abgeglichen(wb["Data_Template"], properties, group_refs)
    dd = DataDictionary(
        source_file=path,
        meta=meta,
        classes=classes,
        properties=properties,
        class_properties=class_properties,
        allowed_values=allowed_values,
        concept_relations=[],
    )
    setattr(dd, "documents", _parse_documents_abgeglichen(wb["Dokumente"], meta))
    return dd


def _parse_matrix(ws, properties: list[DDProperty]) -> list[DDClassProperty]:
    object_labels = [(_str(v) or "") for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    object_ids = [(_str(v) or "") for v in next(ws.iter_rows(min_row=5, max_row=5, values_only=True))]

    obj_cols: list[tuple[int, str]] = []
    for idx, obj_id in enumerate(object_ids, start=1):
        if idx >= 8 and obj_id and obj_id != "Objekt-ID":
            obj_cols.append((idx, obj_id))

    prop_by_en = {p.name_en: p for p in properties if p.name_en}
    prop_by_de = {p.name_de: p for p in properties if p.name_de}
    cps: list[DDClassProperty] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
        vals = list(row)
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        pset = _str(vals[2])
        merkmal = _str(vals[3])
        property_en = _str(vals[4])
        if not (pset or merkmal or property_en):
            continue
        prop = (property_en and prop_by_en.get(property_en)) or (merkmal and prop_by_de.get(merkmal))
        if not prop:
            continue
        for col_idx, class_code in obj_cols:
            cell = _str(vals[col_idx - 1]) if col_idx - 1 < len(vals) else None
            if not cell:
                continue
            marker = cell.strip()
            lower = marker.lower()
            is_assignment = lower == 'x' or marker != ''
            if not is_assignment:
                continue
            override = None if lower == 'x' else marker
            cps.append(DDClassProperty(
                class_code=class_code,
                property_code=prop.code,
                property_set_name=pset or prop.property_set_name,
                is_required=True,
                is_writable=True,
                allowed_values_override=override,
            ))
    return cps


def load_he_dd_v01(path: Path) -> DataDictionary:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)

    if {"Objekte", "Merkmale", "Werte", "Dokumente", "Data_Template", "Dictionary_core", "Dictionary_public"}.issubset(set(wb.sheetnames)) and {"Merkmalgruppen"}.issubset(set(wb.sheetnames)):
        return load_he_dd_abgeglichen(path)

    if {"Objekte", "Werte", "Data Template AreaMgmt", "Dokumente_Dokumentgruppen"}.issubset(set(wb.sheetnames)):
        return load_he_dd_v20260619(path)

    meta = _parse_dictionary_if_present(wb, path)
    classes = _parse_classes(wb["Klassen"], meta)
    classes = _reconcile_class_codes_from_matrix(wb["KlassenMerkmal"], classes)
    properties, allowed_values = _parse_properties(wb["Merkmale_Merkmalsgruppen"], meta)
    class_properties = _parse_matrix(wb["KlassenMerkmal"], properties)
    concept_relations: list[DDConceptRelation] = []
    if "ConceptRelation" in wb.sheetnames:
        ws = wb["ConceptRelation"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            cc = _str(row[0]) if len(row) > 0 else None
            ct = _str(row[1]) if len(row) > 1 else None
            rt = _str(row[2]) if len(row) > 2 else None
            ru = _str(row[3]) if len(row) > 3 else None
            notes = _str(row[4]) if len(row) > 4 else None
            if cc and rt and ru:
                concept_relations.append(DDConceptRelation(cc, ct, rt, ru, notes))

    dd = DataDictionary(
        source_file=path,
        meta=meta,
        classes=classes,
        properties=properties,
        class_properties=class_properties,
        allowed_values=allowed_values,
        concept_relations=concept_relations,
    )
    setattr(dd, "documents", _parse_documents(wb["Dokumente_Dokumentgruppen"], meta) if "Dokumente_Dokumentgruppen" in wb.sheetnames else [])
    return dd
