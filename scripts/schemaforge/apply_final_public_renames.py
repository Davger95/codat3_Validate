from openpyxl import load_workbook
from pathlib import Path

path = Path('/home/Dave/.openclaw/workspace-datadict/P_workspace_codat3_validator/templates/Strukturvorlage_DataDictionary_v2_empty.xlsx')
wb = load_workbook(path)

# Header: remove manual sample defaults so empty template remains truly empty in required fields
ws = wb['Header']
for r in [4,5,6,8,9,10,13]:
    ws.cell(r,2).value = None

# Keep guidance notes/requirements only
ws.cell(10,3).value = 'SYSTEM-GENERATED'
ws.cell(10,4).value = 'Derived from OrganizationCode and DictionaryCode; canonical base pattern https://example.com/<OrganizationCode>/<DictionaryCode>'

# Classes rename
ws = wb['Classes']
ws.cell(1,31).value = 'RelatedDocumentName (EN)'
ws.cell(5,31).value = 'referenced'
ws.cell(7,31).value = 'must match a DocumentName (EN) entry from Documents.DocumentName (EN)'

# Properties rename
ws = wb['Properties']
ws.cell(1,17).value = 'EnumerationDesignation (EN)'
ws.cell(5,17).value = 'referenced'
ws.cell(7,17).value = 'must match a Designation (EN) entry from Values.Designation (EN)'

# Documents rename and remove column O
ws = wb['Documents']
ws.cell(1,4).value = 'GUID/URI'
ws.cell(1,11).value = 'Security level/Sicherheitsstufe/Niveau de sécurité/Livello di sicurezza'
ws.cell(1,12).value = 'Accessibility/Zugänglichkeit/Accessibilité/Accessibilità'
ws.cell(7,18).value = 'date format 2026-06-18T15:30+02:00'

# Delete whole column O if still present as empty spacer
ws.delete_cols(15, 1)

# Re-apply expected governance/status/version/prov headers after delete shift
# After deleting O, Governance->O, Status->P, Version date->Q, Provenance->R
ws.cell(1,15).value = 'Governance'
ws.cell(1,16).value = 'Status'
ws.cell(1,17).value = 'Version date'
ws.cell(1,18).value = 'Provenance (PROV)'
ws.cell(6,16).value = 'Status'
ws.cell(7,16).value = 'must match Rules.Status'
ws.cell(7,17).value = 'date format 2026-06-18T15:30+02:00'
ws.cell(7,15).value = None

# Update comments/notes in version date columns if present as cell comments anywhere in workbook
for sheet_name, col in [('Classes', 28), ('Properties', 29), ('Values', 17), ('Documents', 17), ('GroupOfProperties', 13)]:
    ws2 = wb[sheet_name]
    for row in range(8, ws2.max_row + 1):
        cell = ws2.cell(row, col)
        if cell.comment and 'YYYY-MM-DD' in cell.comment.text:
            cell.comment.text = cell.comment.text.replace('YYYY-MM-DD', '2026-06-18T15:30+02:00')
        if cell.comment and 'Versionsdatum Datum im Format YYYY-MM-DD eintrage.' in cell.comment.text:
            cell.comment.text = cell.comment.text.replace('Versionsdatum Datum im Format YYYY-MM-DD eintrage.', 'Version date as ISO 8601 date-time with timezone, e.g. 2026-06-18T15:30+02:00.')

wb.save(path)
print('final public renames applied')
