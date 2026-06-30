from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

path = Path('/home/Dave/.openclaw/workspace-datadict/P_workspace_codat3_validator/templates/Strukturvorlage_DataDictionary_v2_empty.xlsx')
wb = load_workbook(path)

rules = wb['Rules']
header = wb['Header']

# Add LifecycleStatus options to Rules column J
rules.cell(1, 10).value = 'LifecycleStatus'
values = ['Preview', 'Active', 'Inactive', 'Deprecated', 'Retired', 'Candidate', 'Recorded', 'Superseded', 'Incomplete']
for idx, value in enumerate(values, start=2):
    rules.cell(idx, 10).value = value

# Add / replace defined name
wb.defined_names['LifecycleStatus'] = DefinedName('LifecycleStatus', attr_text="'Rules'!$J$2:$J$10")

# Remove existing validation on B13 if any overlaps, then add proper dropdown
existing = getattr(header.data_validations, 'dataValidation', [])
kept = []
for dv in existing:
    if 'B13' in str(dv.sqref):
        continue
    kept.append(dv)
header.data_validations.dataValidation = kept

dv = DataValidation(type='list', formula1='=LifecycleStatus', allow_blank=True)
dv.prompt = 'Choose an allowed LifecycleStatus value.'
dv.promptTitle = 'LifecycleStatus'
dv.error = 'Please choose a value from the LifecycleStatus list.'
dv.errorTitle = 'Invalid LifecycleStatus'
header.add_data_validation(dv)
dv.add('B13')

wb.save(path)
print('header lifecycle dropdown added')
