from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

XLSX = Path('/home/Dave/.openclaw/workspace-datadict/P_workspace_codat3_validator/templates/Strukturvorlage_DataDictionary_v2_empty.xlsx')
EXPECTED_SHEETS = ['Header','Dictionary_public','Classes','Properties','Values','Documents','GroupOfProperties','Rules','Data_Template']
RULE_NAMES = {
    'Objekt_Einordnung': "'Rules'!$A$2:$A$47",
    'DataType': "'Rules'!$B$2:$B$47",
    'IFC_Data_Type': "'Rules'!$C$2:$C$47",
    'Category': "'Rules'!$D$2:$D$47",
    'Base_Type': "'Rules'!$E$2:$E$47",
    'Beschreibung': "'Rules'!$F$2:$F$47",
    'Status': "'Rules'!$G$2:$G$47",
    'ISG___Informationssicherheitsgesetz': "'Rules'!$H$2:$H$47",
    'FAIR_Prinzipien': "'Rules'!$I$2:$I$47",
}
OLD_SHEET_NAMES = ['Dropdownregeln','Objekte','Merkmale','Werte','Dokumente','Merkmalgruppen','Dictionary_core']


def rebuild_defined_names():
    wb = load_workbook(XLSX)
    for sheet in EXPECTED_SHEETS:
        if sheet not in wb.sheetnames:
            raise RuntimeError(f'missing expected sheet: {sheet}')

    keep = []
    for name, defn in wb.defined_names.items():
        text = getattr(defn, 'attr_text', '') or ''
        if '[' in text:
            continue
        if any(old in text for old in OLD_SHEET_NAMES):
            continue
        if name in RULE_NAMES:
            continue
        if '#REF!' in text and not name.startswith('_xlnm._FilterDatabase'):
            continue
        keep.append((name, defn))

    wb.defined_names = wb.defined_names.__class__()
    for name, defn in keep:
        wb.defined_names[name] = defn
    for name, ref in RULE_NAMES.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

    for ws in wb.worksheets:
        dvs = getattr(ws.data_validations, 'dataValidation', [])
        for dv in dvs:
            if isinstance(dv.formula1, str):
                dv.formula1 = dv.formula1.replace('Dropdownregeln', 'Rules')
            if isinstance(dv.formula2, str):
                dv.formula2 = dv.formula2.replace('Dropdownregeln', 'Rules')
    wb.save(XLSX)


def strip_external_links_zip_level():
    tmpdir = Path(tempfile.mkdtemp(prefix='xlsxfix_'))
    out = tmpdir / XLSX.name
    with ZipFile(XLSX, 'r') as zin, ZipFile(out, 'w', ZIP_DEFLATED) as zout:
        workbook_rels_root = None
        content_types_root = None
        workbook_xml_root = None
        ns_ct = {'ct': 'http://schemas.openxmlformats.org/package/2006/content-types'}
        ns_wb = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main', 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}

        for item in zin.infolist():
            name = item.filename
            if name.startswith('xl/externalLinks/'):
                continue
            if name == 'xl/_rels/workbook.xml.rels':
                root = ET.fromstring(zin.read(name))
                for rel in list(root):
                    typ = rel.attrib.get('Type', '')
                    target = rel.attrib.get('Target', '')
                    if 'externalLink' in typ or 'externalLinks/' in target:
                        root.remove(rel)
                workbook_rels_root = root
                continue
            if name == '[Content_Types].xml':
                root = ET.fromstring(zin.read(name))
                for child in list(root):
                    part = child.attrib.get('PartName', '')
                    ctype = child.attrib.get('ContentType', '')
                    if 'externalLinks' in part or 'externalLink' in ctype:
                        root.remove(child)
                content_types_root = root
                continue
            if name == 'xl/workbook.xml':
                root = ET.fromstring(zin.read(name))
                extrefs = root.find('m:externalReferences', ns_wb)
                if extrefs is not None:
                    root.remove(extrefs)
                defined_names = root.find('m:definedNames', ns_wb)
                if defined_names is not None:
                    for dn in list(defined_names):
                        text = dn.text or ''
                        dn_name = dn.attrib.get('name', '')
                        if '[' in text or 'Dropdownregeln' in text or '#REF!' in text and dn_name not in RULE_NAMES:
                            defined_names.remove(dn)
                workbook_xml_root = root
                continue
            zout.writestr(item, zin.read(name))

        if content_types_root is not None:
            zout.writestr('[Content_Types].xml', ET.tostring(content_types_root, encoding='utf-8', xml_declaration=True))
        if workbook_rels_root is not None:
            zout.writestr('xl/_rels/workbook.xml.rels', ET.tostring(workbook_rels_root, encoding='utf-8', xml_declaration=True))
        if workbook_xml_root is not None:
            zout.writestr('xl/workbook.xml', ET.tostring(workbook_xml_root, encoding='utf-8', xml_declaration=True))
    shutil.move(out, XLSX)
    shutil.rmtree(tmpdir)


def verify():
    with ZipFile(XLSX) as z:
        names = z.namelist()
        assert not any(n.startswith('xl/externalLinks/') for n in names)
        wb_rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='ignore')
        assert 'externalLink' not in wb_rels
        workbook_xml = z.read('xl/workbook.xml').decode('utf-8', errors='ignore')
        assert 'externalReferences' not in workbook_xml
        assert 'Dropdownregeln' not in workbook_xml
        assert '[' not in workbook_xml

    wb = load_workbook(XLSX)
    for sheet in EXPECTED_SHEETS:
        assert sheet in wb.sheetnames
    for name, defn in wb.defined_names.items():
        text = getattr(defn, 'attr_text', '') or ''
        assert '[' not in text, (name, text)
        assert 'Dropdownregeln' not in text, (name, text)
        assert all(old not in text for old in OLD_SHEET_NAMES), (name, text)
    for ws in wb.worksheets:
        dvs = getattr(ws.data_validations, 'dataValidation', [])
        for dv in dvs:
            for f in [dv.formula1, dv.formula2]:
                if isinstance(f, str):
                    assert '[' not in f, (ws.title, f)
                    assert 'Dropdownregeln' not in f, (ws.title, f)
    print('verification ok')


if __name__ == '__main__':
    rebuild_defined_names()
    strip_external_links_zip_level()
    verify()
