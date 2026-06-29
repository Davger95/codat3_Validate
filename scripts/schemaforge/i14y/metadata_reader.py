"""
SchemaForge i14y Pipeline — metadata_reader.py
===============================================
Reads the Dictionary core / Dictionary public metadata tabs from the DD Excel file
and returns a typed DictionaryMetadata dataclass.

This is the single source of truth for all i14y metadata —
no hardcoding elsewhere in the pipeline.

Dictionary tab field → DCAT-AP CH mapping is documented inline.
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import json
import openpyxl


@dataclass
class DictionaryMetadata:
    """All metadata needed to build a DCAT-AP CH dataset entry on i14y."""

    # ── Identification ────────────────────────────────────────────────────────
    organization_code: str        # e.g. "kbob"
    dictionary_code: str          # e.g. "dd-kbob" → becomes i14y identifier
    title_de: str                 # dct:title@de
    title_fr: Optional[str]       # dct:title@fr
    title_en: Optional[str]       # dct:title@en
    version: str                  # dcat:version  e.g. "0.3.0"
    dictionary_uri: str           # base IRI

    # ── Governance ────────────────────────────────────────────────────────────
    publisher_name: str           # foaf:Agent name (must match i14y publisher registry)
    contact_email: Optional[str]  # vcard:hasEmail on dcat:contactPoint
    lifecycle_status: str         # e.g. "Preview" → maps to registrationStatus
    release_date: Optional[str]   # dct:issued  (YYYY-MM-DD)
    qa_procedure: Optional[str]   # dct:conformsTo label
    qa_procedure_url: Optional[str] # dct:conformsTo URI

    # ── Licensing ─────────────────────────────────────────────────────────────
    license_id: Optional[str]     # e.g. "CC-BY-4.0"
    license_url: Optional[str]    # URI to license

    # ── Language ──────────────────────────────────────────────────────────────
    primary_language: str         # e.g. "de-CH"

    # ── Swiss data ecosystem / optional fields ───────────────────────────────
    ech_theme: Optional[str] = None      # dcat:theme IRI or code
    sparql_endpoint: Optional[str] = None
    ttl_download_url: Optional[str] = None
    json_download_url: Optional[str] = None
    more_info_url: Optional[str] = None
    access_rights: Optional[str] = None

    # ── Standards alignment ───────────────────────────────────────────────────
    conforms_to: list[str] = field(default_factory=list)  # list of URIs
    metadata_languages: list[str] = field(default_factory=list)
    spatial_coverage: list[str] = field(default_factory=list)
    temporal_coverage: list[str] = field(default_factory=list)

    # ── Description ───────────────────────────────────────────────────────────
    description_de: Optional[str] = None
    description_fr: Optional[str] = None
    description_en: Optional[str] = None

    # ── Keywords (from Excel) ─────────────────────────────────────────────────
    keywords_de: list[str] = field(default_factory=list)
    keywords_fr: list[str] = field(default_factory=list)
    keywords_en: list[str] = field(default_factory=list)
    version_notes_de: Optional[str] = None
    version_notes_fr: Optional[str] = None
    version_notes_en: Optional[str] = None
    documentation_urls: list[str] = field(default_factory=list)
    related_resource_urls: list[str] = field(default_factory=list)
    applicable_legislation_uris: list[str] = field(default_factory=list)

    # ── Source file tracking ──────────────────────────────────────────────────
    source_file: Optional[str] = None  # absolute path to the Excel file


_PLACEHOLDER = "34"  # Sentinel used in the Excel template for unfilled fields


def _clean(value) -> Optional[str]:
    """Return None for empty or placeholder values."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == _PLACEHOLDER:
        return None
    return s


def _split_multi(value: Optional[str]) -> list[str]:
    if not value:
        return []
    raw = str(value).strip()
    if not raw:
        return []
    if raw.startswith('[') and raw.endswith(']'):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass
    parts = []
    for token in raw.replace(';', '\n').splitlines():
        token = token.strip()
        if token:
            parts.append(token)
    return parts


def _read_sheet_data(ws) -> dict[str, Optional[str]]:
    data: dict[str, Optional[str]] = {}
    for row in ws.iter_rows(values_only=True):
        field_name = _clean(row[0]) if row else None
        if not field_name:
            continue
        if field_name.startswith("──") or field_name.startswith("LEGEND"):
            continue
        if field_name in ("Field", "REQUIRED", "RECOMMENDED", "OPTIONAL"):
            continue
        value = _clean(row[1]) if len(row) > 1 else None
        data[field_name] = value
    return data


def read_metadata(excel_path: str | Path) -> DictionaryMetadata:
    """
    Read Dictionary metadata tabs from the DD Excel and return DictionaryMetadata.

    Preferred structure:
      - 'Dictionary core'   → mandatory minimal authoring block
      - 'Dictionary public' → optional extended publication block

    Backward compatibility:
      - if only 'Dictionary' exists, read from that sheet.

    Expects rows of the form:
        (FieldName, Value, ...)

    Skips section headers (rows where Value is None and FieldName starts with '──').
    Unknown fields are silently ignored.
    """
    excel_path = Path(excel_path).resolve()
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    if "Dictionary_core" in wb.sheetnames:
        core_sheet = "Dictionary_core"
        public_sheet = "Dictionary_public" if "Dictionary_public" in wb.sheetnames else None
    elif "Dictionary" in wb.sheetnames:
        core_sheet = "Dictionary"
        public_sheet = None
    else:
        raise ValueError(f"No 'Dictionary core' or legacy 'Dictionary' sheet in {excel_path}")

    core_data = _read_sheet_data(wb[core_sheet])
    public_data = _read_sheet_data(wb[public_sheet]) if public_sheet else {}
    data = {**core_data, **public_data}

    wb.close()

    # ── Collect conformsTo URIs ───────────────────────────────────────────────
    conforms_to = []
    for key in ("ConformsTo_ISO23386", "ConformsTo_ISO12006_3", "ConformsTo_DCAT_AP_CH"):
        uri = data.get(key)
        if uri:
            conforms_to.append(uri)

    # ── Keywords: prefer explicit Excel metadata, otherwise fallback ─────────
    keywords_de = _split_multi(data.get("Keywords (DE)")) or ["Elementplan", "Data Dictionary", "BIM", "IFC", "FM", "KBOB"]
    keywords_fr = _split_multi(data.get("Keywords (FR)"))
    keywords_en = _split_multi(data.get("Keywords (EN)"))

    # ── Build and validate ────────────────────────────────────────────────────
    title_de = data.get("DictionaryName (DE)")
    dictionary_code = data.get("DictionaryCode")
    version = data.get("DictionaryVersion")

    missing = []
    if not title_de:
        missing.append("DictionaryName (DE)")
    if not dictionary_code:
        missing.append("DictionaryCode")
    if not version:
        missing.append("DictionaryVersion")
    if missing:
        raise ValueError(
            f"Required metadata fields missing or empty in {excel_path}: {missing}"
        )

    return DictionaryMetadata(
        organization_code=data.get("OrganizationCode") or "kbob",
        dictionary_code=dictionary_code,
        title_de=title_de,
        title_fr=data.get("DictionaryName (FR)"),
        title_en=data.get("DictionaryName (EN)"),
        version=version,
        dictionary_uri=data.get("DictionaryUri") or "https://www.kbob.admin.ch",
        publisher_name=data.get("Owner / Publisher") or "KBOB",
        contact_email=data.get("ContactEmail"),
        lifecycle_status=data.get("LifecycleStatus") or "Preview",
        release_date=data.get("ReleaseDate"),
        qa_procedure=data.get("QualityAssuranceProcedure"),
        qa_procedure_url=data.get("QualityAssuranceProcedureUrl"),
        license_id=data.get("License"),
        license_url=data.get("LicenseUrl"),
        primary_language=data.get("PrimaryLanguage") or "de",
        metadata_languages=_split_multi(data.get("MetadataLanguages")) or [data.get("PrimaryLanguage") or "de"],
        ech_theme=data.get("eCH_Theme"),
        access_rights=data.get("AccessRights"),
        spatial_coverage=_split_multi(data.get("SpatialCoverage")),
        temporal_coverage=_split_multi(data.get("TemporalCoverage")),
        sparql_endpoint=data.get("SPARQL_Endpoint"),
        ttl_download_url=data.get("TTL_Download_URL"),
        json_download_url=data.get("JSON_Download_URL"),
        more_info_url=data.get("MoreInfoUrl"),
        conforms_to=conforms_to,
        description_de=data.get("Description (DE)"),
        description_fr=data.get("Description (FR)"),
        description_en=data.get("Description (EN)"),
        keywords_de=keywords_de,
        keywords_fr=keywords_fr,
        keywords_en=keywords_en,
        version_notes_de=data.get("VersionNotes (DE)"),
        version_notes_fr=data.get("VersionNotes (FR)"),
        version_notes_en=data.get("VersionNotes (EN)"),
        documentation_urls=_split_multi(data.get("DocumentationURLs")),
        related_resource_urls=_split_multi(data.get("RelatedResourceURLs")),
        applicable_legislation_uris=_split_multi(data.get("ApplicableLegislation")),
        source_file=str(excel_path),
    )


def validate_metadata(meta: DictionaryMetadata) -> list[str]:
    """
    Return a list of human-readable warnings for fields that are missing
    but recommended for a complete i14y entry.

    Does NOT raise — the caller decides whether to abort or proceed with warnings.
    """
    warnings = []
    if not meta.description_de:
        warnings.append("Description (DE) is missing — required if Dictionary public is used for i14y publication")
    if not meta.description_fr:
        warnings.append("Description (FR) is missing — recommended for Swiss federal datasets")
    if not meta.contact_email:
        warnings.append("ContactEmail is missing — required if Dictionary public is used for i14y publication")
    if not meta.sparql_endpoint and not meta.ttl_download_url:
        warnings.append("No distribution URL (SPARQL_Endpoint or TTL_Download_URL) — dataset will have no distribution")
    if not meta.release_date:
        warnings.append("ReleaseDate is missing — recommended (dct:issued)")
    return warnings
