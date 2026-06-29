from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urlparse

import openpyxl
import rdflib

try:
    from .readers.excel_reader import load_dd
except ImportError:
    from readers.excel_reader import load_dd

WORKSPACE = Path('/home/Dave/.openclaw/workspace-datadict')
BSDD_TTL = Path('/home/Dave/.openclaw/shared/ontologies/bsdd/ifc4.3-bsdd-harvested-official-api.ttl.tmp')
BSDD_URI_CACHE = Path('/home/Dave/.openclaw/shared/ontologies/bsdd/ifc4.3-uri-cache.json')
QUDT_UNITS_TTL = Path('/home/Dave/.openclaw/shared/ontologies/qudt/units.ttl')
VALID_BSDD_URI_PREFIXES = (
    'https://identifier.buildingsmart.org/uri/buildingsmart/ifc/',
)
REQUIRED_SHEETS = [
    'Dictionary_core',
    'Klassen',
    'Merkmale',
    'Dokumente',
    'KlassenMerkmal',
]
REQUIRED_SHEETS_V20260619 = [
    'Objekte',
    'Merkmale',
    'Werte',
    'Dokumente',
    'Data Template AreaMgmt',
]
REQUIRED_SHEETS_ABGEGLICHEN = [
    'Dictionary_core',
    'Dictionary_public',
    'Objekte',
    'Merkmale',
    'Werte',
    'Dokumente',
    'Merkmalgruppen',
    'Data_Template',
]
OPTIONAL_SHEETS = ['ConceptRelation']
ALLOWED_LIFECYCLE = {'Preview', 'Active', 'Inactive', 'Deprecated', 'Retired', 'Candidate', 'Recorded', 'Superseded', 'Incomplete'}
ALLOWED_BASE_TYPES = {'STRING', 'INTEGER', 'REAL', 'BOOLEAN', 'TIME', 'DATETIME'}
ALLOWED_REL_TYPES = {
    'skos:exactMatch',
    'skos:closeMatch',
    'skos:narrowMatch',
    'skos:broadMatch',
    'skos:relatedMatch',
    'owl:sameAs',
}
ALLOWED_CONCEPT_TYPES = {'Class', 'Property', 'AllowedValue', 'PropertySet', 'Enumeration', 'Other'}
SEMVER_RE = re.compile(r'^[0-9]+\.[0-9]+\.[0-9]+$')


@dataclass
class Finding:
    level: str
    code: str
    message: str
    sheet: str | None = None
    cell: str | None = None
    row: int | None = None


class Validator:
    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        self.findings: list[Finding] = []
        self.normalizations: list[dict] = []
        self.wb = openpyxl.load_workbook(workbook_path, data_only=True)
        self.dd = None
        self.ifc_uri_set = None
        self.qudt_unit_labels = None
        self.dd_loaded = False
        self.ifc_uri_set_loaded = False

    @staticmethod
    def _row_has_meaningful_content(row) -> bool:
        meaningful = []
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            meaningful.append(s)
        if not meaningful:
            return False
        joined = ' '.join(meaningful).lower()
        guidance_markers = [
            'should reference',
            'should match',
            'fill if',
            'leave empty',
            'required human label',
            'required canonical',
            'validator should',
            'reference-based validation target',
            'list-based validation',
            'object columns: x = assigned',
            'override values must match',
            'system generated',
            'optional external',
            'required registered source code',
            'required; valid',
            'required de label',
            'required en/canonical label',
            'recommended de definition',
            'recommended en definition',
            'official ifc property uri',
            'official ifc set reference',
            'guid/uri: if applicable',
            'objekt-id: system generated',
            'bezeichnung: required',
            'designation: required',
            'document identification: optional',
            'dokument uri: optional',
            'dokument name: required',
            'revision: recommended',
            'owner: recommended',
            'documentcode: optional',
        ]
        if any(marker in joined for marker in guidance_markers):
            return False
        if len(meaningful) == 1:
            token = meaningful[0].lower()
            if token.startswith('http://') or token.startswith('https://'):
                return False
        return True

    def _iter_data_rows(self, ws, min_row: int):
        for idx, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
            if not self._row_has_meaningful_content(row):
                continue
            meaningful = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if meaningful and all(v.startswith('http://') or v.startswith('https://') for v in meaningful):
                continue
            yield idx, row

    def get_dd(self):
        if self.dd is None:
            self.dd = load_dd(self.workbook_path)
            self.dd_loaded = True
        return self.dd

    def get_ifc_uri_set(self) -> set[str]:
        if self.ifc_uri_set is None:
            self.ifc_uri_set = self._load_ifc_uri_set()
            self.ifc_uri_set_loaded = True
        return self.ifc_uri_set

    def get_qudt_unit_labels(self) -> dict[str, dict[str, set[str]]]:
        if self.qudt_unit_labels is None:
            self.qudt_unit_labels = self._load_qudt_unit_labels()
        return self.qudt_unit_labels

    @staticmethod
    def _normalize_predefined_type(value: str | None) -> str | None:
        if not value:
            return None
        normalized = str(value).strip().replace(' ', '').replace('-', '_')
        return normalized.upper() if normalized else None

    def _extract_ifc_class_from_uri(self, ifc_uri: str | None) -> str | None:
        if not ifc_uri or '/class/' not in ifc_uri:
            return None
        tail = ifc_uri.split('/class/', 1)[1].strip()
        return self._extract_base_ifc_entity(tail)

    def _extract_base_ifc_entity(self, ifc_entity: str | None) -> str | None:
        if not ifc_entity or not str(ifc_entity).strip().startswith('Ifc'):
            return None
        candidate = str(ifc_entity).strip()
        ifc_uri_set = self.get_ifc_uri_set()
        if candidate_uri := f'https://identifier.buildingsmart.org/uri/buildingsmart/ifc/4.3/class/{candidate}':
            if not ifc_uri_set or candidate_uri in ifc_uri_set:
                return candidate
        normalized = candidate.upper()
        for uri in ifc_uri_set:
            if '/class/' not in uri:
                continue
            tail = uri.split('/class/', 1)[1]
            if tail.upper() == normalized and tail.startswith('Ifc'):
                return tail
            if tail.startswith(candidate) and len(tail) > len(candidate):
                suffix = tail[len(candidate):]
                if suffix == suffix.upper() and any(ch.isalpha() for ch in suffix):
                    return candidate
        m = re.match(r'^(Ifc[A-Za-z]+)', candidate)
        return m.group(1) if m else candidate

    def validate_predefined_type(self, ifc_obj: str | None, predefined: str | None, ifc_uri: str | None, sheet_name: str, row_idx: int):
        if not predefined:
            return
        normalized_predefined = self._normalize_predefined_type(predefined)
        if not normalized_predefined:
            return
        ifc_uri_set = self.get_ifc_uri_set()
        if not ifc_obj:
            self.add('warning', 'predefined_type_without_ifc_entity', f'PredefinedType is filled but IfcObject Entity is missing, so authoritative validation cannot be completed: {predefined}', sheet=sheet_name, row=row_idx)
            return
        base_ifc_obj = self._extract_base_ifc_entity(ifc_obj) or ifc_obj
        candidate_uri = f'https://identifier.buildingsmart.org/uri/buildingsmart/ifc/4.3/class/{base_ifc_obj}{normalized_predefined}'
        if ifc_uri_set and candidate_uri not in ifc_uri_set:
            self.add('error', 'invalid_predefined_type', f'PredefinedType {predefined} is not a valid authoritative IFC predefined type for {base_ifc_obj}', sheet=sheet_name, row=row_idx)
            return
        derived_ifc_class = self._extract_ifc_class_from_uri(ifc_uri)
        if derived_ifc_class and derived_ifc_class != base_ifc_obj:
            self.add('error', 'ifc_uri_entity_mismatch', f'IFC URI implies base IFC class {derived_ifc_class}, but IfcObject Entity is {ifc_obj}', sheet=sheet_name, row=row_idx)
            return
        if ifc_uri and ifc_uri != candidate_uri:
            self.add('error', 'ifc_uri_predefined_mismatch', f'IFC URI {ifc_uri} does not match the authoritative IFC predefined-type URI implied by IfcObject Entity + PredefinedType: {candidate_uri}', sheet=sheet_name, row=row_idx)

    def add(self, level: str, code: str, message: str, sheet: str | None = None, cell: str | None = None, row: int | None = None):
        self.findings.append(Finding(level, code, message, sheet, cell, row))

    def add_normalization(self, sheet: str, row: int | None, column: str | None, original_value, normalized_value, reason: str, normalization_kind: str, safe_auto_normalization: bool = True):
        self.normalizations.append({
            'sheet': sheet,
            'row': row,
            'column': column,
            'original_value': original_value,
            'normalized_value': normalized_value,
            'reason': reason,
            'normalization_kind': normalization_kind,
            'safe_auto_normalization': safe_auto_normalization,
        })

    @staticmethod
    def _normalize_label(value: str | None) -> str | None:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        return re.sub(r'\s+', ' ', s).strip().casefold()

    def _load_qudt_unit_labels(self) -> dict[str, dict[str, set[str]]]:
        labels_by_uri: dict[str, dict[str, set[str]]] = {}
        if not QUDT_UNITS_TTL.exists():
            self.add('warning', 'qudt_reference_missing', f'QUDT units ontology not found: {QUDT_UNITS_TTL}')
            return labels_by_uri
        try:
            g = rdflib.Graph()
            g.parse(QUDT_UNITS_TTL, format='turtle')
            query = """
            PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
            SELECT ?unit ?label WHERE {
              ?unit rdfs:label ?label .
            }
            """
            for row in g.query(query):
                uri = str(row.unit)
                label = row.label
                lang = (label.language or '').lower()
                if lang.startswith('en'):
                    lang_key = 'en'
                elif lang.startswith('de'):
                    lang_key = 'de'
                elif lang.startswith('fr'):
                    lang_key = 'fr'
                elif lang.startswith('it'):
                    lang_key = 'it'
                else:
                    continue
                labels_by_uri.setdefault(uri, {}).setdefault(lang_key, set()).add(self._normalize_label(str(label)))
        except Exception as e:
            self.add('warning', 'qudt_reference_load_failed', f'Failed to load QUDT units ontology {QUDT_UNITS_TTL}: {e}')
        return labels_by_uri

    def _load_ifc_uri_set(self) -> set[str]:
        if BSDD_URI_CACHE.exists():
            try:
                payload = json.loads(BSDD_URI_CACHE.read_text())
                return set(payload.get('uris', []))
            except Exception as e:
                self.add('warning', 'ifc_cache_load_failed', f'Failed to load IFC URI cache {BSDD_URI_CACHE}: {e}')
        if not BSDD_TTL.exists():
            self.add('warning', 'ifc_reference_missing', f'Authoritative IFC reference TTL not found: {BSDD_TTL}')
            return set()
        uri_re = re.compile(r'https://identifier\.buildingsmart\.org/uri/buildingsmart/ifc/4\.3(?:\.0)?/[^\s<>"]+')
        uris = set()
        with BSDD_TTL.open('r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                for m in uri_re.findall(line):
                    uris.add(m)
        return uris

    def validate(self):
        self.validate_required_sheets()
        self.validate_dictionary()
        self.validate_wertekatalog()
        self.validate_merkmalsgruppenkatalog()
        self.validate_klassen()
        self.validate_properties()
        self.validate_documents()
        self.validate_matrix()
        self.validate_concept_relations()
        self.validate_pipeline_minimums()
        return self.build_report()

    def is_v20260619_template(self) -> bool:
        return {'Objekte', 'Werte', 'Data Template AreaMgmt'}.issubset(set(self.wb.sheetnames)) and 'Merkmale' in self.wb.sheetnames

    def is_abgeglichen_template(self) -> bool:
        return {'Dictionary_core', 'Dictionary_public', 'Objekte', 'Merkmale', 'Werte', 'Dokumente', 'Data_Template', 'Merkmalgruppen'}.issubset(set(self.wb.sheetnames))

    def validate_required_sheets(self):
        if self.is_abgeglichen_template():
            required = REQUIRED_SHEETS_ABGEGLICHEN
        elif self.is_v20260619_template():
            required = REQUIRED_SHEETS_V20260619
        elif 'Dictionary_core' in self.wb.sheetnames or 'Dictionary_public' in self.wb.sheetnames:
            required = REQUIRED_SHEETS
        else:
            required = [
                'Dictionary',
                'Klassen',
                'Merkmale',
                'Dokumente',
                'KlassenMerkmal',
            ]
        missing = [s for s in required if s not in self.wb.sheetnames]
        for sheet in missing:
            self.add('error', 'missing_sheet', f'Missing required sheet: {sheet}', sheet=sheet)

    def _dictionary_sheet_names(self) -> tuple[str | None, str | None]:
        core = 'Dictionary_core' if 'Dictionary_core' in self.wb.sheetnames else ('Dictionary' if 'Dictionary' in self.wb.sheetnames else None)
        public = 'Dictionary_public' if 'Dictionary_public' in self.wb.sheetnames else None
        return core, public

    def _dict_rows(self, sheet_name: str) -> dict[str, tuple[str | None, int]]:
        ws = self.wb[sheet_name]
        rows = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            field = row[0] if len(row) > 0 else None
            value = row[1] if len(row) > 1 else None
            if field is not None:
                rows[str(field).strip()] = (None if value is None else str(value).strip(), i)
        return rows

    def validate_dictionary(self):
        core_sheet, public_sheet = self._dictionary_sheet_names()
        if not core_sheet:
            return
        core_rows = self._dict_rows(core_sheet)
        public_rows = self._dict_rows(public_sheet) if public_sheet else {}

        core_required = [
            'OrganizationCode',
            'DictionaryCode',
            'DictionaryName (DE)',
            'DictionaryVersion',
            'DictionaryUri',
            'LifecycleStatus',
        ]
        for key in core_required:
            value_row = core_rows.get(key)
            if not value_row or not value_row[0]:
                self.add('error', 'missing_dictionary_field', f'Missing required dictionary core value: {key}', sheet=core_sheet, row=value_row[1] if value_row else None)

        status = (core_rows.get('LifecycleStatus') or [None])[0]
        if status and status not in ALLOWED_LIFECYCLE:
            self.add('error', 'invalid_lifecycle', f'Invalid LifecycleStatus: {status}', sheet=core_sheet)
        version = (core_rows.get('DictionaryVersion') or [None])[0]
        if version and not SEMVER_RE.match(version):
            self.add('error', 'invalid_semver', f'DictionaryVersion must be semantic version, got: {version}', sheet=core_sheet)
        uri = (core_rows.get('DictionaryUri') or [None])[0]
        if uri and not self.is_absolute_uri(uri):
            self.add('error', 'invalid_uri', f'DictionaryUri is not a valid absolute IRI: {uri}', sheet=core_sheet)

        if public_sheet:
            contact_email = (public_rows.get('ContactEmail') or [None])[0]
            if contact_email and '@' not in contact_email:
                self.add('warning', 'invalid_contact_email_public', f'ContactEmail in Dictionary_public is invalid: {contact_email}', sheet=public_sheet)
            primary_language = (public_rows.get('PrimaryLanguage') or [None])[0]
            if primary_language and not re.match(r'^[a-z]{2}(?:-[A-Z]{2})?$', primary_language):
                self.add('warning', 'invalid_primary_language_public', f'PrimaryLanguage in Dictionary_public should be ISO-like language tag (e.g. de or de-CH), got: {primary_language}', sheet=public_sheet)
            release_date = (public_rows.get('ReleaseDate') or [None])[0]
            if release_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', release_date):
                self.add('warning', 'invalid_release_date_public', f'ReleaseDate in Dictionary_public should be YYYY-MM-DD, got: {release_date}', sheet=public_sheet)
            modified_date = (public_rows.get('ModifiedDate') or [None])[0]
            if modified_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', modified_date):
                self.add('warning', 'invalid_modified_date_public', f'ModifiedDate in Dictionary_public should be YYYY-MM-DD, got: {modified_date}', sheet=public_sheet)

    def is_valid_bsdd_identifier_uri(self, value: str) -> bool:
        return any(value.startswith(prefix) for prefix in VALID_BSDD_URI_PREFIXES)

    @staticmethod
    def slugify(value: str | None) -> str:
        if not value:
            return ''
        s = str(value).strip().lower()
        s = re.sub(r'[^a-z0-9]+', '-', s)
        s = re.sub(r'-+', '-', s).strip('-')
        return s

    def parse_allowed_list(self, raw: str | None, sheet: str | None = None, row: int | None = None, column: str | None = None) -> list[str]:
        if not raw:
            return []
        txt = str(raw).strip()
        if txt.startswith('[') and txt.endswith(']'):
            try:
                parsed = json.loads(txt)
                if isinstance(parsed, list):
                    return [str(x).strip() for x in parsed if str(x).strip()]
            except Exception:
                pass
        parts = [p.strip().strip('"').strip("'") for p in re.split(r'[;,]', txt) if p and p.strip()]
        normalized = [p for p in parts if p]
        normalized_json = json.dumps(normalized, ensure_ascii=False)
        if txt != normalized_json and sheet:
            self.add_normalization(sheet, row, column, txt, normalized_json, 'Legacy delimiter-based list normalized to JSON-style list for validation comparison', 'list-format-normalization', True)
        return normalized

    def is_strict_string_list_syntax(self, raw: str | None) -> bool:
        if not raw:
            return False
        txt = str(raw).strip()
        if not (txt.startswith('[') and txt.endswith(']')):
            return False
        try:
            parsed = json.loads(txt)
        except Exception:
            return False
        return isinstance(parsed, list)

    def _require_en_plus_one_local(self, row, idx: int, sheet_name: str, en_idx: int, local_idx_map: dict[str, int], label: str):
        en_value = self._cell(row, en_idx)
        local_values = {lang: self._cell(row, col_idx) for lang, col_idx in local_idx_map.items()}
        if not en_value:
            self.add('error', 'missing_required_english_translation', f'{label} must be filled in English (EN).', sheet=sheet_name, row=idx)
        if not any(local_values.values()):
            langs = '/'.join(local_idx_map.keys())
            self.add('error', 'missing_required_local_translation', f'{label} requires at least one local-language value in {langs} in addition to English.', sheet=sheet_name, row=idx)

    def _header_index_map(self, ws, header_row: int) -> dict[str, int]:
        mapping = {}
        for idx, cell in enumerate(ws[header_row], start=1):
            value = None if cell.value is None else str(cell.value).strip()
            if value:
                mapping[value] = idx
        return mapping


    def _load_dropdown_values(self, target_label: str) -> set[str]:
        sheet_name = 'Dropdownregeln'
        if sheet_name not in self.wb.sheetnames:
            return set()
        ws = self.wb[sheet_name]

        # First support the current aligned workbook layout where row 1 contains
        # direct column headers like "Status" or "ISG - Informationssicherheitsgesetz"
        # and the allowed values are listed vertically below each header.
        header_map = self._header_index_map(ws, 1)
        normalized_target = target_label.removeprefix('Dropdownregeln.').strip()
        if normalized_target in header_map:
            col_idx = header_map[normalized_target]
            collected = set()
            for row_idx in range(2, ws.max_row + 1):
                raw = ws.cell(row_idx, col_idx).value
                value = None if raw is None else str(raw).strip()
                if value:
                    collected.add(value)
            if collected:
                return collected

        # Backward-compatible fallback for older vertically-sectioned layouts.
        collected = set()
        current_target = None
        for row in ws.iter_rows(values_only=True):
            values = [None if v is None else str(v).strip() for v in row]
            nonempty = [v for v in values if v]
            if not nonempty:
                continue
            first = nonempty[0]
            if first.startswith('Dropdownregeln.'):
                current_target = first
                continue
            if current_target == target_label:
                for v in nonempty:
                    if v != current_target and not v.startswith('Dropdownregeln.'):
                        collected.add(v)
        return collected

    def validate_wertekatalog(self):
        if 'Werte' not in self.wb.sheetnames:
            return
        ws = self.wb['Werte']
        headers = self._header_index_map(ws, 3)
        seen_ids = set()
        for idx, row in self._iter_data_rows(ws, 10):
            katalog_id = self._cell(row, headers.get('Werteliste-ID', headers.get('Werte-ID', 5)))
            label_de = self._cell(row, headers.get('Bezeichnung', 6))
            label_fr = self._cell(row, headers.get('Désignation (FR)', 7))
            label_it = self._cell(row, headers.get('Designazione (IT)', 8))
            label_en = self._cell(row, headers.get('Label (EN)', 9))
            values_en_raw = self._cell(row, headers.get('EnumerationValues\n(EN)', 10))
            values_de_raw = self._cell(row, headers.get('Werteliste\n(DE)', 11))
            values_fr_raw = self._cell(row, headers.get('Liste de valeurs\n(FR)', 12))
            values_it_raw = self._cell(row, headers.get('Lista valori\n(IT)', 13))
            unit = self._cell(row, headers.get('Einheiten', 14))
            status = self._cell(row, headers.get('Status', 16))
            version_date = self._cell(row, headers.get('Versionsdatum', 17))

            if not any([katalog_id, label_de, label_fr, label_it, label_en, values_en_raw, values_de_raw, values_fr_raw, values_it_raw, unit, status, version_date]):
                continue

            if not katalog_id:
                self.add('error', 'missing_value_catalog_id', 'Werte row missing Werteliste-ID.', sheet='Werte', row=idx)
            elif katalog_id in seen_ids:
                self.add('error', 'duplicate_value_catalog_id', f'Duplicate Werteliste-ID: {katalog_id}', sheet='Werte', row=idx)
            else:
                seen_ids.add(katalog_id)

            expected_id = f"{self.slugify(label_en).replace('-', '_')}_enum" if label_en else None
            if katalog_id and expected_id and katalog_id != expected_id:
                self.add('warning', 'noncanonical_value_list_id', f'Werteliste-ID is present but differs from canonical generated form {expected_id}', sheet='Werte', row=idx)
            if not label_en and katalog_id:
                self.add('error', 'missing_required_english_translation', 'Werte.Designation (EN) must be filled in English (EN) if a Werteliste-ID is given.', sheet='Werte', row=idx)
            if not values_en_raw:
                self.add('error', 'missing_required_english_translation', 'Werte.EnumerationValues must be filled in English (EN).', sheet='Werte', row=idx)
            if not any([values_de_raw, values_fr_raw, values_it_raw]):
                self.add('error', 'missing_required_local_translation', 'Werte value lists require at least one local-language value in DE/IT/FR in addition to English.', sheet='Werte', row=idx)

            if label_en and not any([label_de, label_fr, label_it]):
                self.add('error', 'missing_required_local_translation', 'Werte.Designation requires at least one local-language value in DE/IT/FR in addition to English.', sheet='Werte', row=idx)

            for column_label, raw_value in [('EnumerationValues (EN)', values_en_raw), ('Werteliste (DE)', values_de_raw), ('Liste de valeurs (FR)', values_fr_raw), ('Lista valori (IT)', values_it_raw)]:
                if raw_value and not self.is_strict_string_list_syntax(raw_value):
                    self.add('error', 'invalid_list_syntax', f'{column_label} must use JSON-style list syntax like ["One", "Two", "Three"]. Use decimal points, not commas, inside numbers.', sheet='Werte', row=idx)

            values_en = self.parse_allowed_list(values_en_raw, sheet='Werte', row=idx, column='EnumerationValues (EN)') if values_en_raw else []
            values_de = self.parse_allowed_list(values_de_raw, sheet='Werte', row=idx, column='Werteliste (DE)') if values_de_raw else []
            values_fr = self.parse_allowed_list(values_fr_raw, sheet='Werte', row=idx, column='Liste de valeurs (FR)') if values_fr_raw else []
            values_it = self.parse_allowed_list(values_it_raw, sheet='Werte', row=idx, column='Lista valori (IT)') if values_it_raw else []

            if values_en and len(values_en) != len(set(v.casefold() for v in values_en)):
                self.add('error', 'duplicate_enumeration_values', f'EnumerationValues (EN) contains duplicate values for Werteliste-ID {katalog_id or idx}', sheet='Werte', row=idx)
            for column_label, parsed in [('Werteliste (DE)', values_de), ('Liste de valeurs (FR)', values_fr), ('Lista valori (IT)', values_it)]:
                if parsed and len(parsed) != len(set(v.casefold() for v in parsed)):
                    self.add('error', 'duplicate_enumeration_values', f'{column_label} contains duplicate values for Werteliste-ID {katalog_id or idx}', sheet='Werte', row=idx)

            en_len = len(values_en)
            for column_label, parsed in [('Werteliste (DE)', values_de), ('Liste de valeurs (FR)', values_fr), ('Lista valori (IT)', values_it)]:
                if parsed and values_en and len(parsed) != en_len:
                    self.add('error', 'misaligned_multilingual_value_list', f'{column_label} has {len(parsed)} values but EnumerationValues (EN) has {en_len} for Werteliste-ID {katalog_id or idx}.', sheet='Werte', row=idx)

            if status and status not in ALLOWED_LIFECYCLE:
                self.add('error', 'invalid_value_catalog_status', f'Werte.Status must be one of the allowed lifecycle values, got: {status}', sheet='Werte', row=idx)
            if version_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', version_date):
                self.add('error', 'invalid_value_catalog_version_date', f'Werte.Versionsdatum should be YYYY-MM-DD, got: {version_date}', sheet='Werte', row=idx)

    def validate_merkmalsgruppenkatalog(self):
        group_sheet = 'Merkmalgruppen' if 'Merkmalgruppen' in self.wb.sheetnames else ('Merkmalgruppen' if 'Merkmalgruppen' in self.wb.sheetnames else None)
        if not group_sheet:
            return
        ws = self.wb[group_sheet]
        headers = self._header_index_map(ws, 2)
        core_sheet, _ = self._dictionary_sheet_names()
        org_code = None
        if core_sheet:
            org_code = (self._dict_rows(core_sheet).get('OrganizationCode') or [None])[0]
        expected_prefix = f"{org_code}_" if org_code else None
        seen_ids = set()
        seen_codes = set()
        seen_labels = {}
        for idx, row in self._iter_data_rows(ws, 8):
            group_id = self._cell(row, headers.get('Merkmalsgruppe-ID', 4))
            group_code = self._cell(row, headers.get('Merkmalsgruppe-Code', 5))
            group_en = self._cell(row, headers.get('Designation (EN)', headers.get('Merkmalsgruppe (EN)', 6)))
            desc_en = self._cell(row, headers.get('Description (EN)', 7))
            label_de = self._cell(row, headers.get('Bezeichnung (DE)', headers.get('Beschreibung (DE)', 8)))
            label_fr = self._cell(row, headers.get('Désignation (FR)', headers.get('Description (FR)', 9)))
            label_it = self._cell(row, headers.get('Designazione (IT)', headers.get('Descrizione (IT)', 10)))
            desc_de = label_de
            desc_fr = label_fr
            desc_it = label_it
            status = self._cell(row, headers.get('Status', 12))
            version_date = self._cell(row, headers.get('Versionsdatum', 13))

            if not any([group_id, group_code, group_en, desc_en, desc_de, desc_fr, desc_it, status, version_date]):
                continue

            if not group_id:
                self.add('error', 'missing_group_id', 'Merkmalgruppen row missing Merkmalsgruppe-ID.', sheet=group_sheet, row=idx)
            elif group_id in seen_ids:
                self.add('error', 'duplicate_group_id', f'Duplicate Merkmalsgruppe-ID: {group_id}', sheet=group_sheet, row=idx)
            else:
                seen_ids.add(group_id)

            if group_code:
                if group_code in seen_codes:
                    self.add('error', 'duplicate_group_code', f'Duplicate Merkmalsgruppe-Code: {group_code}', sheet=group_sheet, row=idx)
                else:
                    seen_codes.add(group_code)

            if not group_en:
                self.add('error', 'missing_group_label_en', 'Designation (EN) is required.', sheet=group_sheet, row=idx)
            else:
                if expected_prefix and not group_en.startswith(expected_prefix):
                    self.add('error', 'invalid_group_label_prefix', f'Designation (EN) must start with {expected_prefix}', sheet=group_sheet, row=idx)
                suffix = group_en[len(expected_prefix):] if expected_prefix and group_en.startswith(expected_prefix) else group_en
                if len(suffix) > 16:
                    self.add('error', 'group_label_too_long', f'Designation (EN) name part after prefix must be max 16 characters, got {len(suffix)}.', sheet=group_sheet, row=idx)
                expected_id = self.slugify(group_en)
                if not group_id and expected_id:
                    self.add('warning', 'missing_group_id', f'Merkmalsgruppe-ID is system-generated and should not be authored manually. Generated value: {expected_id}', sheet=group_sheet, row=idx)
                    self.add_normalization(group_sheet, idx, 'Merkmalsgruppe-ID', group_id, expected_id, 'System-generated ID derived from Designation (EN)', 'derived-group-id', True)
                elif group_id and expected_id and group_id != expected_id:
                    self.add('warning', 'system_generated_group_id_override', f'Merkmalsgruppe-ID is a system-generated field. Manual value {group_id} will be overwritten by generated value {expected_id}.', sheet=group_sheet, row=idx)
                    self.add_normalization(group_sheet, idx, 'Merkmalsgruppe-ID', group_id, expected_id, 'Manual ID overridden by system-generated ID derived from Designation (EN)', 'derived-group-id', True)

            for lang, value in [('EN', group_en), ('DE', label_de), ('FR', label_fr), ('IT', label_it)]:
                norm = self._norm(value)
                if norm:
                    seen_labels.setdefault(norm, []).append((lang, idx))
            if not desc_en:
                self.add('error', 'missing_group_description_en', 'Description (EN) is required.', sheet=group_sheet, row=idx)
            if not any([desc_de, desc_fr, desc_it]):
                self.add('error', 'missing_group_description_local', 'At least one local-language designation/description must be filled in DE/FR/IT.', sheet=group_sheet, row=idx)

            if status and status not in ALLOWED_LIFECYCLE:
                self.add('error', 'invalid_group_status', f'Merkmalgruppen.Status must be one of the allowed lifecycle values, got: {status}', sheet=group_sheet, row=idx)
            if version_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', version_date):
                self.add('error', 'invalid_group_version_date', f'Merkmalgruppen.Versionsdatum should be YYYY-MM-DD, got: {version_date}', sheet=group_sheet, row=idx)
        self._check_unique_labels(seen_labels, group_sheet, 'duplicate_group_label', 'Property-group label')

    def _check_unique_labels(self, registry: dict[str, list[tuple[str, int]]], sheet_name: str, error_code: str, label_name: str):
        for norm_value, hits in registry.items():
            if len(hits) > 1:
                refs = ", ".join(f"{lang}@row{row}" for lang, row in hits)
                for lang, row in hits:
                    self.add('error', error_code, f'{label_name} must be unique within {sheet_name}. Duplicate label across multilingual reference fields: {norm_value} ({refs})', sheet=sheet_name, row=row)

    def validate_klassen(self):
        sheet_name = 'Objekte' if ('Objekte' in self.wb.sheetnames and (self.is_v20260619_template() or self.is_abgeglichen_template())) else 'Klassen'
        if sheet_name not in self.wb.sheetnames:
            return
        ws = self.wb[sheet_name]
        headers = self._header_index_map(ws, 3 if self.is_abgeglichen_template() else 3)
        seen_codes = set()
        seen_labels = {}
        dd = self.get_dd()
        ifc_uri_set = self.get_ifc_uri_set()
        document_source_codes = {d.get('SourceCode') for d in getattr(dd, 'documents', []) if d.get('SourceCode')}
        document_ids = {d.get('Document Identification') or d.get('Dokument-ID') or d.get('Document-ID') for d in getattr(dd, 'documents', []) if (d.get('Document Identification') or d.get('Dokument-ID') or d.get('Document-ID'))}
        object_class_allowed = self._load_dropdown_values('Dropdownregeln.Objekt-Einordnung')
        status_allowed = self._load_dropdown_values('Dropdownregeln.Status')
        start_row = 10
        for idx, row in self._iter_data_rows(ws, start_row):
            if self.is_v20260619_template():
                obj_id = self._cell(row, 9)
                obj_einordnung = self._cell(row, 10)
                bezeichnung = self._cell(row, 11)
                designation = self._cell(row, 12)
                beschreibung = self._cell(row, 13)
                description = self._cell(row, 14)
                ifc_uri = self._cell(row, 16)
                ifc_obj = self._cell(row, 17)
                ifc_type = self._cell(row, 18)
                predefined = None
                source = None
                identification = None
                final_name = None
            else:
                if self.is_abgeglichen_template():
                    obj_id = self._cell(row, headers.get('Objekt-ID', 9))
                    obj_einordnung = self._cell(row, headers.get('Objekt-Einordnung ', headers.get('Objekt-Einordnung', 10)))
                    bezeichnung = self._cell(row, headers.get('Bezeichnung', 11))
                    designation = self._cell(row, headers.get('Designation (EN)', headers.get('Designation', 12)))
                    beschreibung = self._cell(row, headers.get('Beschreibung', 15))
                    description = self._cell(row, headers.get('Description', 16))
                    ifc_uri = self._cell(row, headers.get('Objekte.IFC_URI', headers.get('IFC_URI', headers.get('GUID/URI_1', 20))))
                    ifc_obj = self._cell(row, headers.get('IfcObject Entity', 21))
                    ifc_type = self._cell(row, headers.get('IfcTypeObject Entity', 22))
                    predefined = self._cell(row, headers.get('PredefinedType', 23))
                    object_type = self._cell(row, headers.get('ObjectType', 24))
                    status = self._cell(row, headers.get('Status', 26))
                    version_date = self._cell(row, headers.get('Versionsdatum', 27))
                    source = self._cell(row, headers.get('Herkunft (PROV)', 28))
                    related_document = self._cell(row, headers.get('Objekte.RelatedDocument (Document-ID)', headers.get('RelatedDocument (Document-ID)', headers.get('RelatedDocument', 30))))
                    identification = None
                    final_name = None
                    self._require_en_plus_one_local(row, idx, sheet_name, headers.get('Designation (EN)', headers.get('Designation', 12)), {'DE': headers.get('Bezeichnung (DE)', headers.get('Bezeichnung', 11)), 'IT': headers.get('Designazione (IT)', headers.get('Designazione (IT)', 14)), 'FR': headers.get('Désignation (FR)', headers.get('Désignation (FR)', 13))}, 'Objekt.Bezeichnung/Designation')
                    self._require_en_plus_one_local(row, idx, sheet_name, headers.get('Description (EN)', headers.get('Description', 16)), {'DE': headers.get('Beschreibung (DE)', headers.get('Beschreibung', 15)), 'IT': headers.get('Descrizione (IT)', headers.get('Beschreibung (IT)', 18)), 'FR': headers.get('Description (FR)', headers.get('Beschreibung (FR)', 17))}, 'Objekt.Beschreibung/Description')
                    label_de = self._cell(row, headers.get('Bezeichnung (DE)', headers.get('Bezeichnung', 11)))
                    label_fr = self._cell(row, headers.get('Désignation (FR)', 13))
                    label_it = self._cell(row, headers.get('Designazione (IT)', 14))
                    for lang, value in [('DE', label_de), ('EN', designation), ('FR', label_fr), ('IT', label_it)]:
                        norm = self._norm(value)
                        if norm:
                            seen_labels.setdefault(norm, []).append((lang, idx))
                    expected_obj_id = self.slugify(designation or bezeichnung) if (designation or bezeichnung) else None
                    if not obj_id and expected_obj_id:
                        self.add('warning', 'missing_object_id', f'Objekt-ID is system-generated and should not be authored manually. Generated value: {expected_obj_id}', sheet=sheet_name, row=idx)
                        self.add_normalization(sheet_name, idx, 'Objekt-ID', obj_id, expected_obj_id, 'System-generated ID derived from Designation/Bezeichnung', 'derived-object-id', True)
                        obj_id = expected_obj_id
                    elif obj_id and expected_obj_id and obj_id != expected_obj_id:
                        self.add('warning', 'system_generated_object_id_override', f'Objekt-ID is a system-generated field. Manual value {obj_id} will be overwritten by generated value {expected_obj_id}.', sheet=sheet_name, row=idx)
                        self.add_normalization(sheet_name, idx, 'Objekt-ID', obj_id, expected_obj_id, 'Manual ID overridden by system-generated ID derived from Designation/Bezeichnung', 'derived-object-id', True)
                    if object_class_allowed and obj_einordnung and obj_einordnung not in object_class_allowed:
                        self.add('error', 'invalid_objekt_einordnung', f'Objekt-Einordnung must come from Dropdownregeln.Objekt-Einordnung. Got: {obj_einordnung}', sheet=sheet_name, row=idx)
                    if status and status_allowed and status not in status_allowed:
                        self.add('error', 'invalid_object_status', f'Objekte.Status must come from Dropdownregeln.Status. Got: {status}', sheet=sheet_name, row=idx)
                    if version_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', version_date):
                        self.add('error', 'invalid_object_version_date', f'Objekte.Versionsdatum should be YYYY-MM-DD, got: {version_date}', sheet=sheet_name, row=idx)
                else:
                    obj_id = self._cell(row, 7)
                    bezeichnung = self._cell(row, 9)
                    designation = self._cell(row, 10)
                    beschreibung = self._cell(row, 11)
                    description = self._cell(row, 12)
                    ifc_uri = self._cell(row, 13)
                    ifc_obj = self._cell(row, 14)
                    ifc_type = self._cell(row, 15)
                    predefined = self._cell(row, 16)
                    source = self._cell(row, 20)
                    identification = self._cell(row, 23)
                    final_name = self._cell(row, 24)
                    obj_einordnung = self._cell(row, 8)
            if not obj_id:
                derived_id = self.slugify(designation or bezeichnung)
                if derived_id:
                    self.add('warning', 'missing_class_code', f'Objekte row missing Objekt-ID; derivable suggested ID: {derived_id}', sheet=sheet_name, row=idx)
                    self.add_normalization('Klassen', idx, 'Objekt-ID', obj_id, derived_id, 'Missing class ID is derivable from Designation/Bezeichnung', 'derived-id-suggestion', True)
                    obj_id = derived_id
                else:
                    self.add('error', 'missing_class_code', 'Objekte row missing Objekt-ID and no derivable Designation/Bezeichnung is available', sheet=sheet_name, row=idx)
            elif obj_id in seen_codes:
                self.add('error', 'duplicate_class_code', f'Duplicate Objekt-ID: {obj_id}', sheet=sheet_name, row=idx)
            else:
                seen_codes.add(obj_id)
            if not (bezeichnung or designation):
                self.add('error', 'missing_class_label', 'Objekte row missing Bezeichnung/Designation', sheet=sheet_name, row=idx)
            if not (beschreibung or description):
                self.add('error', 'missing_class_definition', 'Objekte row missing Beschreibung/Description', sheet=sheet_name, row=idx)
            if not obj_einordnung:
                self.add('warning', 'missing_objekt_einordnung', 'Objekte row missing Objekt-Einordnung', sheet=sheet_name, row=idx)
            if not ifc_uri:
                self.add('error', 'missing_ifc_uri', 'Objekte row missing IFC URI', sheet=sheet_name, row=idx)
            else:
                if not self.is_absolute_uri(ifc_uri):
                    self.add('error', 'invalid_ifc_uri', f'Invalid IFC URI: {ifc_uri}', sheet=sheet_name, row=idx)
                elif not self.is_valid_bsdd_identifier_uri(ifc_uri):
                    self.add('error', 'invalid_ifc_uri_namespace', f'IFC URI is not in a valid buildingSMART/bSDD identifier namespace: {ifc_uri}', sheet=sheet_name, row=idx)
                elif ifc_uri_set and ifc_uri not in ifc_uri_set:
                    self.add('error', 'unknown_ifc_uri', f'IFC URI not found in authoritative bSDD harvest: {ifc_uri}', sheet=sheet_name, row=idx)
            if not ifc_obj:
                self.add('error', 'missing_ifc_object_entity', 'Objekte row missing IfcObject Entity', sheet=sheet_name, row=idx)
            if not ifc_type:
                self.add('warning', 'missing_ifc_type_object_entity', 'IfcTypeObject Entity is missing; this may be acceptable if mapping exists only on object level', sheet=sheet_name, row=idx)
            self.validate_predefined_type(ifc_obj, predefined, ifc_uri, sheet_name, idx)
            if not self.is_v20260619_template():
                classification_in_use = bool(source or identification or final_name)
                if self.is_abgeglichen_template():
                    if not source:
                        self.add('error', 'missing_prov_source', 'Objekte.Herkunft (PROV) is required and may be user-defined.', sheet=sheet_name, row=idx)
                    if not related_document:
                        self.add_normalization(sheet_name, idx, 'Objekte.RelatedDocument (Document-ID)', related_document, 'Organisation', 'Empty RelatedDocument defaults to Organisation', 'default-related-document', True)
                    elif document_ids and related_document not in document_ids:
                        self.add('error', 'unknown_related_document_id', f'Objekte.RelatedDocument (Document-ID) must reference an existing Dokumente.Document-ID. Got: {related_document}', sheet=sheet_name, row=idx)
                    if ifc_obj and (not self._extract_base_ifc_entity(ifc_obj) or (ifc_uri_set and f'https://identifier.buildingsmart.org/uri/buildingsmart/ifc/4.3/class/{self._extract_base_ifc_entity(ifc_obj)}' not in ifc_uri_set)):
                        self.add('error', 'invalid_ifc_object_entity', f'IfcObject Entity must be a valid IFC entity. Got: {ifc_obj}', sheet=sheet_name, row=idx)
                    if ifc_type and (not self._extract_base_ifc_entity(ifc_type) or (ifc_uri_set and f'https://identifier.buildingsmart.org/uri/buildingsmart/ifc/4.3/class/{self._extract_base_ifc_entity(ifc_type)}' not in ifc_uri_set)):
                        self.add('error', 'invalid_ifc_type_object_entity', f'IfcTypeObject Entity must be a valid IFC entity when filled. Got: {ifc_type}', sheet=sheet_name, row=idx)
                    if object_type and (not self._extract_base_ifc_entity(object_type) or (ifc_uri_set and f'https://identifier.buildingsmart.org/uri/buildingsmart/ifc/4.3/class/{self._extract_base_ifc_entity(object_type)}' not in ifc_uri_set)):
                        self.add('error', 'invalid_object_type', f'ObjectType must be a valid IFC entity. Got: {object_type}', sheet=sheet_name, row=idx)
                else:
                    if not source:
                        self.add('error', 'missing_source', 'Objekte.Herkunft (PROV) is required and may be user-defined.', sheet=sheet_name, row=idx)
                    elif source != 'Organisation' and document_source_codes and source not in document_source_codes:
                        self.add('error', 'unknown_source_code', f'Klassen.Source not registered in Dokumente.SourceCode: {source}. Add the source formally or use Organisation.', sheet=sheet_name, row=idx)
                if classification_in_use and source and source != 'Organisation' and not identification:
                    self.add('warning', 'missing_identification', 'Classification source is given but Identification is missing. Please add it if applicable, otherwise ignore.', sheet=sheet_name, row=idx)
                if classification_in_use and (source or identification) and not final_name:
                    self.add('warning', 'missing_classification_name', 'Classification section is in use but final Name column is missing.', sheet=sheet_name, row=idx)
        if self.is_abgeglichen_template():
            self._check_unique_labels(seen_labels, sheet_name, 'duplicate_object_label', 'Object label')

    def validate_properties(self):
        sheet_name = 'Merkmale' if self.is_abgeglichen_template() and 'Merkmale' in self.wb.sheetnames else 'Merkmale'
        if sheet_name not in self.wb.sheetnames:
            return
        ws = self.wb[sheet_name]
        headers = self._header_index_map(ws, 2 if self.is_abgeglichen_template() else 3)
        ifc_uri_set = self.get_ifc_uri_set()
        seen_codes = set()
        seen_labels = {}
        seen_prop_short_codes = set()
        if (self.is_v20260619_template() or self.is_abgeglichen_template()) and 'Werte' in self.wb.sheetnames:
            value_ids = set()
            vws = self.wb['Werte']
            for vr in vws.iter_rows(min_row=10, values_only=True):
                vid = self._cell(vr, 5)
                if vid:
                    value_ids.add(vid)
        else:
            value_ids = set()
        start_row = 8 if self.is_v20260619_template() else (8 if self.is_abgeglichen_template() else 4)
        for idx, row in self._iter_data_rows(ws, start_row):
            if self.is_v20260619_template():
                prop_code = self._cell(row, 5)
                merkmal = self._cell(row, 6)
                prop_en = self._cell(row, 7)
                data_type = self._cell(row, 10)
                data_type_ifc = self._cell(row, 11)
                value_list_id = self._cell(row, 12)
                value_list = None
                ifc_property_uri = self._cell(row, 14)
                ifc_pset = self._cell(row, 15)
                ifc_qto = self._cell(row, 16)
                custom_pset = self._cell(row, 17)
                local_group = self._cell(row, 17)
            else:
                if self.is_abgeglichen_template():
                    prop_id = self._cell(row, headers.get('Merkmal-ID', 5))
                    prop_code = prop_id or self._cell(row, headers.get('Merkmal-Code', 6))
                    merkmal = self._cell(row, headers.get('Bezeichnung (DE)', headers.get('Merkmal', 7)))
                    prop_en = self._cell(row, headers.get('Designation (EN)', headers.get('Property', 8)))
                    data_type = self._cell(row, headers.get('DataType\n(Base Type)', 13))
                    data_type_ifc = self._cell(row, headers.get('DataType\n(IFC)', 14))
                    value_list_id = self._cell(row, headers.get('Werteliste-ID', 17))
                    value_list = None
                    ifc_property_uri = self._cell(row, headers.get('IFC_URI', headers.get('GUID/URI_1', 19)))
                    ifc_pset = self._cell(row, headers.get('IfcPropertySet (Pset)\nIfcQuantitySet (Qto)', 20))
                    ifc_qto = None
                    custom_pset = None
                    local_group = self._cell(row, 2)
                    status = self._cell(row, headers.get('Status', 28))
                    version_date = self._cell(row, headers.get('Versionsdatum', 29))
                    self._require_en_plus_one_local(row, idx, sheet_name, headers.get('Designation (EN)', headers.get('Property', 8)), {'DE': headers.get('Bezeichnung (DE)', headers.get('Merkmal', 7)), 'IT': headers.get('Designazione (IT)', headers.get('IT', 10)), 'FR': headers.get('Désignation (FR)', headers.get('FR', 9))}, 'Merkmale.Bezeichnung/Designation')
                    self._require_en_plus_one_local(row, idx, sheet_name, headers.get('Description (EN)', 12), {'DE': headers.get('Beschreibung (DE)', headers.get('Beschreibung', 11)), 'IT': headers.get('Descrizione (IT)', headers.get('Beschreibung (IT)', 14)), 'FR': headers.get('Description (FR)', headers.get('Beschreibung (FR)', 13))}, 'Merkmale.Beschreibung/Description')
                    label_fr = self._cell(row, headers.get('Désignation (FR)', headers.get('FR', 9)))
                    label_it = self._cell(row, headers.get('Designazione (IT)', headers.get('IT', 10)))
                    for lang, value in [('DE', merkmal), ('EN', prop_en), ('FR', label_fr), ('IT', label_it)]:
                        norm = self._norm(value)
                        if norm:
                            seen_labels.setdefault(norm, []).append((lang, idx))
                else:
                    prop_id = None
                    prop_code = self._cell(row, 4)
                    merkmal = self._cell(row, 5)
                    prop_en = self._cell(row, 6)
                    data_type = self._cell(row, 9)
                    data_type_ifc = self._cell(row, 10)
                    value_list_id = self._cell(row, 11)
                    value_list = self._cell(row, 12)
                    ifc_property_uri = self._cell(row, 14)
                    ifc_pset = self._cell(row, 15)
                    ifc_qto = self._cell(row, 16)
                    custom_pset = self._cell(row, 17)
                    local_group = self._cell(row, 20)
            expected_prop_id = self.slugify(prop_en) if prop_en else None
            if self.is_abgeglichen_template():
                if not prop_id:
                    if expected_prop_id:
                        self.add('warning', 'missing_property_code', f'Merkmal-ID is system-generated and should not be authored manually. Generated value: {expected_prop_id}', sheet=sheet_name, row=idx)
                        self.add_normalization(sheet_name, idx, 'Merkmal-ID', prop_id, expected_prop_id, 'System-generated ID derived from Property (EN)', 'derived-property-id', True)
                        prop_code = expected_prop_id
                    else:
                        self.add('error', 'missing_property_code', 'Property row missing Merkmal-ID and no usable Property (EN) is available for generation.', sheet=sheet_name, row=idx)
                elif expected_prop_id and prop_id != expected_prop_id:
                    self.add('warning', 'system_generated_property_id_override', f'Merkmal-ID is a system-generated field. Manual value {prop_id} will be overwritten by generated value {expected_prop_id}.', sheet=sheet_name, row=idx)
                    self.add_normalization(sheet_name, idx, 'Merkmal-ID', prop_id, expected_prop_id, 'Manual ID overridden by system-generated ID derived from Property (EN)', 'derived-property-id', True)
                    prop_code = expected_prop_id
            if not prop_code:
                self.add('error', 'missing_property_code', 'Property row missing Merkmal-ID', sheet=sheet_name, row=idx)
            elif prop_code in seen_codes:
                self.add('error', 'duplicate_property_code', f'Duplicate Merkmal-ID: {prop_code}', sheet=sheet_name, row=idx)
            else:
                seen_codes.add(prop_code)
            prop_short_code = self._cell(row, headers.get('Merkmal-Code', 6)) if self.is_abgeglichen_template() else None
            if prop_short_code:
                if prop_short_code in seen_prop_short_codes:
                    self.add('error', 'duplicate_property_short_code', f'Duplicate Merkmal-Code: {prop_short_code}', sheet=sheet_name, row=idx)
                else:
                    seen_prop_short_codes.add(prop_short_code)
            if not merkmal:
                self.add('error', 'missing_property_label_de', 'Property row missing Merkmal', sheet=sheet_name, row=idx)
            if not prop_en:
                self.add('error', 'missing_property_label_en', 'Property row missing Property', sheet=sheet_name, row=idx)
            if not data_type:
                self.add('error', 'missing_data_type', 'Property row missing DataType (Base Type)', sheet=sheet_name, row=idx)
            elif data_type.upper() not in ALLOWED_BASE_TYPES:
                self.add('error', 'invalid_data_type', f'Invalid DataType (Base Type): {data_type}', sheet=sheet_name, row=idx)
            if not data_type_ifc:
                self.add('error', 'missing_ifc_data_type', 'Property row missing DataType (IFC)', sheet=sheet_name, row=idx)
            unit_code = self._cell(row, headers.get('Einheit-Code', 21)) if self.is_abgeglichen_template() else None
            unit_qudt = self._cell(row, headers.get('QUDT URI', 22)) if self.is_abgeglichen_template() else None
            unit_name_de = self._cell(row, headers.get('Einheit-Name (DE)', 23)) if self.is_abgeglichen_template() else None
            unit_name_fr = self._cell(row, headers.get('Einheit-Name (FR)', 24)) if self.is_abgeglichen_template() else None
            unit_name_it = self._cell(row, headers.get('Einheit-Name (IT)', 25)) if self.is_abgeglichen_template() else None
            unit_name_en = self._cell(row, headers.get('Einheit-Name (EN)', 26)) if self.is_abgeglichen_template() else None
            if value_list_id and self.is_v20260619_template() and value_ids:
                if value_list_id not in value_ids:
                    self.add('error', 'unknown_value_list_id', f'Werteliste-ID {value_list_id} is not registered in Werte.', sheet=sheet_name, row=idx)
            if ifc_property_uri:
                if not self.is_absolute_uri(ifc_property_uri):
                    self.add('error', 'invalid_ifc_property_uri', f'IFC_URI is not a valid absolute IRI: {ifc_property_uri}', sheet=sheet_name, row=idx)
                elif not self.is_valid_bsdd_identifier_uri(ifc_property_uri):
                    self.add('error', 'invalid_ifc_property_uri_namespace', f'IFC_URI is not in a valid buildingSMART/bSDD identifier namespace: {ifc_property_uri}', sheet=sheet_name, row=idx)
                elif ifc_uri_set and ifc_property_uri not in ifc_uri_set:
                    self.add('error', 'unknown_ifc_property_uri', f'IFC_URI not found in authoritative bSDD harvest: {ifc_property_uri}', sheet=sheet_name, row=idx)
            for uri_label, uri_value in [('IfcPropertySet / IfcQuantitySet', ifc_pset or ifc_qto)]:
                if uri_value:
                    if not ifc_property_uri:
                        self.add('warning', 'ifc_set_without_ifc_uri', f'{uri_label} is filled but Merkmale.IFC_URI is empty. The set reference is only validated when IFC_URI is also provided.', sheet=sheet_name, row=idx)
                    elif not self.is_absolute_uri(uri_value):
                        self.add('error', 'invalid_ifc_linked_uri', f'{uri_label} is not a valid absolute IRI: {uri_value}', sheet=sheet_name, row=idx)
                    elif not self.is_valid_bsdd_identifier_uri(uri_value):
                        self.add('error', 'invalid_ifc_linked_uri_namespace', f'{uri_label} is not in a valid buildingSMART/bSDD identifier namespace: {uri_value}', sheet=sheet_name, row=idx)
                    elif ifc_uri_set and uri_value not in ifc_uri_set:
                        self.add('error', 'unknown_ifc_linked_uri', f'{uri_label} not found in authoritative bSDD harvest: {uri_value}', sheet=sheet_name, row=idx)
            expected_value_list_id = f"{self.slugify(prop_en or merkmal).replace('-', '_')}_enum" if (prop_en or merkmal) else None
            if value_list_id:
                if expected_value_list_id and value_list_id != expected_value_list_id:
                    self.add('warning', 'noncanonical_value_list_id', f'Werteliste-ID is present but differs from canonical generated form {expected_value_list_id}', sheet=sheet_name, row=idx)
            elif value_list or ((self.is_v20260619_template() or self.is_abgeglichen_template()) and value_ids):
                self.add('warning', 'missing_value_list_id', f'Werteliste-ID is missing; derivable generated ID: {expected_value_list_id}', sheet=sheet_name, row=idx)
                if expected_value_list_id:
                    self.add_normalization(sheet_name, idx, 'Werteliste-ID', value_list_id, expected_value_list_id, 'Missing Werteliste-ID is derivable from Property/Merkmal', 'derived-enumeration-id', True)
            if value_list:
                parsed = self.parse_allowed_list(value_list, sheet=sheet_name, row=idx, column='Werteliste')
                if len(parsed) != len(set(parsed)):
                    self.add('error', 'duplicate_enumeration_values', f'Werteliste contains duplicate values for {prop_code}', sheet=sheet_name, row=idx)
            if self.is_abgeglichen_template():
                if status and status not in ALLOWED_LIFECYCLE:
                    self.add('error', 'invalid_property_status', f'Merkmale.Status must be one of the allowed lifecycle values, got: {status}', sheet=sheet_name, row=idx)
                if version_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', version_date):
                    self.add('error', 'invalid_property_version_date', f'Merkmale.Versionsdatum should be YYYY-MM-DD, got: {version_date}', sheet=sheet_name, row=idx)
            if not unit_qudt:
                self.add('error', 'missing_qudt_unit_uri', 'QUDT URI must be filled for Merkmale rows.', sheet=sheet_name, row=idx)
            else:
                if not self.is_absolute_uri(unit_qudt):
                    self.add('error', 'invalid_qudt_unit_uri', f'QUDT URI is not a valid absolute IRI: {unit_qudt}', sheet=sheet_name, row=idx)
                elif not unit_qudt.startswith('http://qudt.org/vocab/unit/') and not unit_qudt.startswith('https://qudt.org/vocab/unit/'):
                    self.add('error', 'invalid_qudt_unit_namespace', f'QUDT URI is not in the QUDT unit namespace: {unit_qudt}', sheet=sheet_name, row=idx)
                else:
                    qudt_labels = self.get_qudt_unit_labels()
                    if unit_qudt not in qudt_labels:
                        self.add('error', 'unknown_qudt_unit_uri', f'QUDT URI not found in local QUDT ontology: {unit_qudt}', sheet=sheet_name, row=idx)
                    else:
                        for lang_key, raw_name, column_label in [
                            ('de', unit_name_de, 'Einheit-Name (DE)'),
                            ('fr', unit_name_fr, 'Einheit-Name (FR)'),
                            ('it', unit_name_it, 'Einheit-Name (IT)'),
                            ('en', unit_name_en, 'Einheit-Name (EN)'),
                        ]:
                            expected = qudt_labels[unit_qudt].get(lang_key, set())
                            canonical = sorted(expected)[0] if expected else None
                            if not raw_name and canonical:
                                self.add('warning', 'missing_qudt_unit_label', f'{column_label} is system-generated from QUDT and should not be authored manually. Generated value: {canonical}', sheet=sheet_name, row=idx)
                                self.add_normalization(sheet_name, idx, column_label, raw_name, canonical, 'System-generated unit label derived from local QUDT rdfs:label', 'derived-qudt-unit-label', True)
                            elif raw_name:
                                normalized = self._normalize_label(raw_name)
                                if normalized not in expected:
                                    self.add('warning', 'system_generated_qudt_unit_label_override', f'{column_label} is a system-generated field. Manual value {raw_name} will be overwritten by ontology-derived value {canonical or "<no label>"}.', sheet=sheet_name, row=idx)
                                    if canonical:
                                        self.add_normalization(sheet_name, idx, column_label, raw_name, canonical, 'Manual unit label overridden by local QUDT rdfs:label', 'derived-qudt-unit-label', True)
        if self.is_abgeglichen_template():
            self._check_unique_labels(seen_labels, sheet_name, 'duplicate_property_label', 'Property label')

    def validate_documents(self):
        dd = self.get_dd()
        docs = getattr(dd, 'documents', [])
        sheet_name = 'Dokumente' if self.is_abgeglichen_template() and 'Dokumente' in self.wb.sheetnames else 'Dokumente'
        doc_start_row = 13
        if sheet_name == 'Dokumente':
            ws = self.wb[sheet_name]
            doc_start_row = 8
            for r in range(1, min(ws.max_row, 20) + 1):
                vals = [None if c.value is None else str(c.value).strip() for c in ws[r]]
                if 'DocumentName (EN)' in vals or 'DocumentationName' in vals or 'DocumentName' in vals:
                    doc_start_row = r + 1
                    break
        seen_doc_ids = set()
        seen_doc_codes = set()
        seen_group_map = {}
        document_security_allowed = self._load_dropdown_values('Dropdownregeln.ISG - Informationssicherheitsgesetz')
        document_access_allowed = self._load_dropdown_values('Dropdownregeln.FAIR Prinzipien')
        status_allowed = self._load_dropdown_values('Dropdownregeln.Status')
        for i, doc in enumerate(docs, start=doc_start_row):
            doc_id = doc.get('Document Identification') or doc.get('Dokument-ID')
            doc_name = doc.get('DocumentName') or doc.get('DocumentationName') or doc.get('Dokument Name')
            doc_name_de = doc.get('DocumentName (DE)')
            doc_name_fr = doc.get('DocumentationName (FR)')
            doc_name_it = doc.get('DocumentationName (IT)')
            revision = doc.get('Revision')
            owner = doc.get('Owner')
            doc_code = doc.get('DocumentCode')
            group_code = doc.get('DocumentGroupCode')
            group_label = doc.get('DocumentGroupName') or doc.get('DocumentGroupLabel')
            status = doc.get('status') or doc.get('Status')
            version_date = doc.get('Versionsdatum') or doc.get('VersionDate')
            if self.is_abgeglichen_template() and not any([doc_name, doc_name_de, doc_name_fr, doc_name_it, revision, owner, doc_code, group_code, group_label, status, version_date]):
                continue
            expected_doc_id = self.slugify(doc_name) if doc_name else None
            if not doc_id and expected_doc_id:
                self.add('warning', 'missing_document_id', f'Document-ID is system-generated and should not be authored manually. Generated value: {expected_doc_id}', sheet=sheet_name, row=i)
                self.add_normalization(sheet_name, i, 'Document-ID', doc_id, expected_doc_id, 'System-generated ID derived from DocumentName', 'derived-document-id', True)
            elif doc_id and expected_doc_id and doc_id != expected_doc_id:
                self.add('warning', 'system_generated_document_id_override', f'Document-ID is a system-generated field. Manual value {doc_id} will be overwritten by generated value {expected_doc_id}.', sheet=sheet_name, row=i)
                self.add_normalization(sheet_name, i, 'Document-ID', doc_id, expected_doc_id, 'Manual ID overridden by system-generated ID derived from DocumentName', 'derived-document-id', True)
            if not doc_name:
                self.add('error', 'missing_document_name', 'Document row missing DocumentName', sheet=sheet_name, row=i)
            if doc_name and not any([doc_name_de, doc_name_fr, doc_name_it]):
                self.add('error', 'missing_required_local_translation', 'DocumentName requires at least one local-language value in DE/IT/FR in addition to English.', sheet=sheet_name, row=i)
            if revision is None or str(revision).strip() == '':
                self.add('error', 'missing_revision', 'Document row missing Revision', sheet=sheet_name, row=i)
            if doc_code:
                if doc_code in seen_doc_codes:
                    self.add('error', 'duplicate_document_code', f'Duplicate Document-Code: {doc_code}', sheet=sheet_name, row=i)
                else:
                    seen_doc_codes.add(doc_code)
            if not group_code:
                self.add('error', 'missing_document_group_code', 'Document row missing Sicherheitsstufe', sheet=sheet_name, row=i)
            elif document_security_allowed and group_code not in document_security_allowed:
                self.add('error', 'invalid_document_security_level', f'Dokumente.Sicherheitsstufe must come from Dropdownregeln.ISG - Informationssicherheitsgesetz. Got: {group_code}', sheet=sheet_name, row=i)
            if not group_label:
                self.add('error', 'missing_document_group_label', 'Document row missing Zugänglichkeit', sheet=sheet_name, row=i)
            elif document_access_allowed and group_label not in document_access_allowed:
                self.add('error', 'invalid_document_accessibility', f'Dokumente.Zugänglichkeit must come from Dropdownregeln.FAIR Prinzipien. Got: {group_label}', sheet=sheet_name, row=i)
            elif group_code:
                prev = seen_group_map.get(group_code)
                if prev and prev != group_label:
                    self.add('error', 'inconsistent_group_label', f'Sicherheitsstufe {group_code} maps to multiple Zugänglichkeit labels: {prev} / {group_label}', sheet=sheet_name, row=i)
                seen_group_map[group_code] = group_label
            if status and status_allowed and status not in status_allowed:
                self.add('error', 'invalid_document_status', f'Dokumente.Status must come from Dropdownregeln.Status. Got: {status}', sheet=sheet_name, row=i)
            if version_date and not re.match(r'^\d{4}-\d{2}-\d{2}$', str(version_date).strip()):
                self.add('error', 'invalid_document_version_date', f'Dokumente.Versionsdatum should be YYYY-MM-DD, got: {version_date}', sheet=sheet_name, row=i)
            doc_uri = doc.get('Dokument URI')
            if doc_uri and not self.is_absolute_uri(doc_uri):
                self.add('error', 'invalid_document_uri', f'Dokument URI is invalid: {doc_uri}', sheet=sheet_name, row=i)

    def validate_matrix(self):
        matrix_sheet = 'Data_Template' if self.is_abgeglichen_template() and 'Data_Template' in self.wb.sheetnames else ('Data Template AreaMgmt' if self.is_v20260619_template() and 'Data Template AreaMgmt' in self.wb.sheetnames else 'KlassenMerkmal')
        if matrix_sheet not in self.wb.sheetnames:
            return
        ws = self.wb[matrix_sheet]
        dd = self.get_dd()
        class_codes = {c.code for c in dd.classes}
        property_name_de = {p.name_de: p.code for p in dd.properties if p.name_de}
        property_name_en = {p.name_en: p.code for p in dd.properties if p.name_en}
        property_codes_registered = {p.code for p in dd.properties if p.code}
        registered_property_sets = set()
        for p in dd.properties:
            if getattr(p, 'ifc_pset_uri', None):
                registered_property_sets.add(getattr(p, 'ifc_pset_uri'))
            if getattr(p, 'property_set_name', None):
                registered_property_sets.add(getattr(p, 'property_set_name'))
            if getattr(p, 'raw_ifc_qto', None):
                registered_property_sets.add(getattr(p, 'raw_ifc_qto'))
            if getattr(p, 'merkmalsgruppe', None):
                registered_property_sets.add(getattr(p, 'merkmalsgruppe'))
        if self.is_v20260619_template():
            property_cols = []
            for col_idx in range(5, ws.max_column + 1):
                prop_code = self._cell([ws.cell(3, col_idx).value], 1)
                if prop_code:
                    property_cols.append((col_idx, prop_code))
                    if prop_code not in property_codes_registered:
                        self.add('error', 'matrix_unknown_merkmal_id', f'Data Template AreaMgmt references unknown Merkmal-ID in row 3: {prop_code}', sheet=matrix_sheet, row=3)
            if not property_cols:
                self.add('error', 'matrix_missing_property_columns', 'Data Template AreaMgmt has no property IDs in row 3 from column 5 onward', sheet=matrix_sheet, row=3)
            for ridx in range(8, ws.max_row + 1):
                row_values = [ws.cell(ridx, c).value for c in range(1, ws.max_column + 1)]
                class_code = self._cell([ws.cell(ridx, 1).value], 1)
                class_label = self._cell([ws.cell(ridx, 2).value], 1)
                if not self._row_has_meaningful_content(row_values):
                    continue
                if not class_code:
                    continue
                if class_code not in class_codes:
                    self.add('error', 'matrix_unknown_class', f'Data Template AreaMgmt row references unknown class/object ID: {class_code}', sheet=matrix_sheet, row=ridx)
                for col_idx, prop_code in property_cols:
                    cell = self._cell([ws.cell(ridx, col_idx).value], 1)
                    if not cell:
                        continue
                    if prop_code not in property_codes_registered:
                        continue
                    if cell.lower() != 'x':
                        allowed = self.allowed_values_for_property(prop_code)
                        overrides = self.parse_allowed_list(cell, sheet=matrix_sheet, row=ridx, column=f'col-{col_idx}')
                        allowed_cmp = {a.strip().casefold() for a in allowed}
                        overrides_cmp = {o.strip().casefold() for o in overrides}
                        if allowed and not overrides_cmp.issubset(allowed_cmp):
                            self.add('error', 'invalid_allowed_values_override', f'This object-specific restriction contains values {overrides} that are not listed in the property\'s official allowed values for {prop_code}. Please correct the override or add the missing value to the property\'s allowed values.', sheet=matrix_sheet, row=ridx)
            return
        if self.is_abgeglichen_template():
            object_label_map = {}
            for c in getattr(dd, 'classes', []):
                for value in [getattr(c, 'name_en', None), getattr(c, 'name_de', None), getattr(c, 'name_fr', None), getattr(c, 'name_it', None)]:
                    norm = self._norm(value)
                    if norm:
                        object_label_map[norm] = c.code
            property_label_map = {}
            property_allowed_values = {}
            for p in getattr(dd, 'properties', []):
                property_allowed_values[p.code] = self.allowed_values_for_property(p.code)
                for value in [getattr(p, 'name_en', None), getattr(p, 'name_de', None), getattr(p, 'name_fr', None), getattr(p, 'name_it', None)]:
                    norm = self._norm(value)
                    if norm:
                        property_label_map[norm] = p.code
            group_sheet = 'Merkmalgruppen' if 'Merkmalgruppen' in self.wb.sheetnames else ('Merkmalgruppen' if 'Merkmalgruppen' in self.wb.sheetnames else None)
            group_label_map = {}
            if group_sheet:
                gws = self.wb[group_sheet]
                for ridx, grow in self._iter_data_rows(gws, 8):
                    for col in [6, 8, 9, 10]:
                        val = self._cell(grow, col)
                        norm = self._norm(val)
                        if norm:
                            group_label_map[norm] = val
            property_cols = []
            for col_idx in range(9, ws.max_column + 1):
                label = self._cell([ws.cell(2, col_idx).value], 1)
                prop_identifier = self._cell([ws.cell(3, col_idx).value], 1)
                if col_idx >= 9 and not prop_identifier:
                    continue
                if not label and not prop_identifier:
                    continue
                prop_code = None
                if prop_identifier and prop_identifier in property_allowed_values:
                    prop_code = prop_identifier
                elif label:
                    norm = self._norm(label)
                    prop_code = property_label_map.get(norm)
                if not prop_code:
                    # In aligned Data_Template, row 3 technical identifiers are authoritative.
                    # If there is a technical ID but it is unknown, report it; otherwise ignore decorative row-2 labels.
                    if prop_identifier:
                        display = label or prop_identifier
                        self.add('error', 'matrix_unknown_property_label', f'Data_Template property reference not found in Merkmale labels/IDs (EN/DE/FR/IT): {display}', sheet=matrix_sheet, row=3)
                    continue
                property_cols.append((col_idx, prop_code, label or prop_identifier))
            for ridx in range(8, ws.max_row + 1):
                row_values = [ws.cell(ridx, c).value for c in range(1, ws.max_column + 1)]
                if not self._row_has_meaningful_content(row_values):
                    continue
                object_label = self._cell([ws.cell(ridx, 1).value], 1)
                property_group_label = self._cell([ws.cell(ridx, 3).value], 1)
                if object_label and self._norm(object_label) not in object_label_map:
                    self.add('error', 'matrix_unknown_object_label', f'Data_Template object reference not found in Objekte labels (EN/DE/FR/IT): {object_label}', sheet=matrix_sheet, row=ridx)
                if property_group_label and self._norm(property_group_label) not in group_label_map:
                    self.add('error', 'matrix_unknown_group_label', f'Data_Template property-group reference not found in Merkmalgruppen labels (EN/DE/FR/IT): {property_group_label}', sheet=matrix_sheet, row=ridx)
                for col_idx, prop_code, label in property_cols:
                    cell = self._cell([ws.cell(ridx, col_idx).value], 1)
                    if cell is None or str(cell).strip() == '':
                        continue
                    if str(cell).strip().lower() == 'x':
                        continue
                    overrides = self.parse_allowed_list(cell, sheet=matrix_sheet, row=ridx, column=f'col-{col_idx}')
                    allowed = property_allowed_values.get(prop_code, [])
                    allowed_cmp = {a.strip().casefold() for a in allowed}
                    overrides_cmp = {str(o).strip().casefold() for o in overrides}
                    if allowed and not overrides_cmp.issubset(allowed_cmp):
                        self.add('error', 'invalid_allowed_values_override', f'Data_Template override {overrides} is not a subset of the registered Werte list for property {label} / {prop_code}.', sheet=matrix_sheet, row=ridx)
            return
        row4 = [self._norm(v) for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        row5 = [self._norm(v) for v in next(ws.iter_rows(min_row=5, max_row=5, values_only=True))]
        object_ids = []
        known_class_refs = 0
        for idx, obj_id in enumerate(row5, start=1):
            if idx >= 8 and obj_id and obj_id != 'Objekt-ID':
                object_ids.append((idx, obj_id))
                if obj_id in class_codes:
                    known_class_refs += 1
        if not object_ids:
            self.add('error', 'matrix_missing_object_columns', 'KlassenMerkmal has no object IDs in row 5 from column 8 onward', sheet='KlassenMerkmal', row=5)
        for ridx, row in self._iter_data_rows(ws, 9):
            vals = list(row)
            matrix_merkmal_id = self._cell(vals, 2)
            pset = self._cell(vals, 3)
            merkmal = self._cell(vals, 4)
            prop_en = self._cell(vals, 5)
            if not (matrix_merkmal_id or pset or merkmal or prop_en):
                continue
            if not matrix_merkmal_id:
                self.add('error', 'matrix_missing_merkmal_id', 'KlassenMerkmal row is missing Merkmal-ID; it must reference a registered property identifier from Merkmale.', sheet='KlassenMerkmal', row=ridx)
                continue
            if matrix_merkmal_id not in property_codes_registered:
                self.add('error', 'matrix_unknown_merkmal_id', f'KlassenMerkmal row references unknown Merkmal-ID: {matrix_merkmal_id}', sheet='KlassenMerkmal', row=ridx)
                continue
            if pset and pset not in registered_property_sets:
                self.add('error', 'unknown_property_set_reference', f'KlassenMerkmal PropertySet is not registered in Merkmale as IfcPropertySet, IfcQuantitySet, or Merkmalsgruppe: {pset}', sheet='KlassenMerkmal', row=ridx)
            prop_code = matrix_merkmal_id
            if prop_en and property_name_en.get(prop_en) and property_name_en.get(prop_en) != prop_code:
                self.add('warning', 'matrix_property_name_mismatch', f'Property (EN) does not match the registered Merkmal-ID {prop_code}', sheet='KlassenMerkmal', row=ridx)
            if merkmal and property_name_de.get(merkmal) and property_name_de.get(merkmal) != prop_code:
                self.add('warning', 'matrix_property_label_mismatch', f'Merkmal (DE) does not match the registered Merkmal-ID {prop_code}', sheet='KlassenMerkmal', row=ridx)
            for col_idx, class_code in object_ids:
                cell = self._cell(vals, col_idx)
                if not cell:
                    continue
                if cell.lower() != 'x':
                    allowed = self.allowed_values_for_property(prop_code)
                    overrides = self.parse_allowed_list(cell, sheet='KlassenMerkmal', row=ridx, column=f'object-col-{col_idx}')
                    allowed_cmp = {a.strip().casefold() for a in allowed}
                    overrides_cmp = {o.strip().casefold() for o in overrides}
                    if allowed and not overrides_cmp.issubset(allowed_cmp):
                        self.add('error', 'invalid_allowed_values_override', f'This class-specific restriction contains values {overrides} that are not listed in the property\'s official allowed values for {prop_code}. Please correct the override or add the missing value to the property\'s allowed values.', sheet='KlassenMerkmal', row=ridx)

    def validate_concept_relations(self):
        dd = self.get_dd()
        for idx, cr in enumerate(dd.concept_relations, start=5):
            if cr.concept_type and cr.concept_type not in ALLOWED_CONCEPT_TYPES:
                self.add('error', 'invalid_concept_type', f'Invalid ConceptType: {cr.concept_type}', sheet='ConceptRelation', row=idx)
            if cr.relation_type not in ALLOWED_REL_TYPES:
                self.add('error', 'invalid_relation_type', f'Invalid RelationType: {cr.relation_type}', sheet='ConceptRelation', row=idx)
            if cr.related_uri and not self.is_absolute_uri(cr.related_uri):
                self.add('error', 'invalid_related_uri', f'Invalid RelatedConceptUri: {cr.related_uri}', sheet='ConceptRelation', row=idx)

    def validate_pipeline_minimums(self):
        dd = self.get_dd()
        # Classes are not universally mandatory for sparse guidance/property-only templates.
        if len(dd.properties) < 1 and len(dd.classes) > 0:
            self.add('error', 'minimum_properties', 'At least one valid property row is required when classes are defined')
        require_assignments = len(dd.classes) > 0 and len(dd.properties) > 0
        require_documents = len(dd.classes) > 0 and (len(dd.properties) > 0 or len(dd.class_properties) > 0)
        if require_assignments and len(dd.class_properties) < 1:
            self.add('error', 'minimum_assignments', 'At least one valid class-property assignment is required when both classes and properties are defined')
        if require_documents and len(getattr(dd, 'documents', [])) < 1:
            self.add('error', 'minimum_documents', 'At least one document row is required when class/property content is being defined for source-governed output')

    def allowed_values_for_property(self, prop_code: str) -> list[str]:
        dd = self.get_dd()
        for p in dd.properties:
            if p.code == prop_code and p.enumeration_values:
                return self.parse_allowed_list(p.enumeration_values)
        return []

    @staticmethod
    def is_absolute_uri(value: str) -> bool:
        try:
            parsed = urlparse(value)
            return bool(parsed.scheme and parsed.netloc)
        except Exception:
            return False

    @staticmethod
    def _cell(row, idx_1based: int):
        if idx_1based - 1 < len(row):
            v = row[idx_1based - 1]
            if v is None:
                return None
            s = str(v).strip()
            return s if s else None
        return None

    @staticmethod
    def _norm(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    def build_report(self):
        errors = [f for f in self.findings if f.level == 'error']
        warnings = [f for f in self.findings if f.level == 'warning']
        parser_valid = len([e for e in errors if e.code in {
            'missing_sheet', 'missing_dictionary_field', 'minimum_classes', 'minimum_properties', 'minimum_assignments'
        }]) == 0
        pipeline_valid = len(errors) == 0
        governance_valid = len(errors) == 0 and len(warnings) == 0
        return {
            'workbook': str(self.workbook_path),
            'summary': {
                'parser_valid': parser_valid,
                'pipeline_valid': pipeline_valid,
                'governance_valid': governance_valid,
                'errors': len(errors),
                'warnings': len(warnings),
                'normalizations': len(self.normalizations),
            },
            'findings': [asdict(f) for f in self.findings],
            'normalizations': self.normalizations,
        }


def _layman_mapping(code: str) -> dict:
    mapping = {
        'ifc_reference_missing': {
            'title': 'IFC reference dataset unavailable',
            'what_it_means': 'Der Validator konnte nicht auf das autoritative IFC/bSDD-Referenzdataset zugreifen, das für URI-Gegenprüfungen verwendet wird. Dies ist ein Umgebungs-/Referenzproblem und kein Fehler in einer Workbook-Zeile.',
            'what_to_do': 'Stellen Sie die IFC reference TTL/cache in der Validator-Umgebung wieder her, falls autoritative URI-Existenzprüfungen erforderlich sind. Falls dies in CI nicht verfügbar ist, behandeln Sie dies als System-/Referenzwarnung und nicht als Korrekturaufgabe auf Sheet-/Row-Ebene.',
            'category': 'Validator environment and reference data',
        },
        'missing_ifc_uri': {
            'title': 'Missing IFC reference',
            'what_it_means': 'Diesem Objekt fehlt die offizielle IFC URI reference.',
            'what_to_do': 'Fügen Sie die korrekte offizielle IFC URI für dieses Objekt hinzu.',
            'category': 'Object definitions',
        },
        'missing_class_definition': {
            'title': 'Missing object definition',
            'what_it_means': 'Dieses Objekt hat noch keine klare Beschreibung/Definition.',
            'what_to_do': 'Fügen Sie eine kurze Definition hinzu, damit Benutzer verstehen, was das Objekt bedeutet.',
            'category': 'Object definitions',
        },
        'unknown_value_list_id': {
            'title': 'Value catalog ID is not registered',
            'what_it_means': 'Ein Merkmal verweist auf eine value catalog ID, die im Tab Werte nicht existiert.',
            'what_to_do': 'Prüfen Sie die Schreibweise der Werteliste-ID oder registrieren Sie den fehlenden value catalog in Werte.',
            'category': 'Value catalog issues',
        },
        'noncanonical_value_list_id': {
            'title': 'Value catalog ID does not follow the standard format',
            'what_it_means': 'Die Werteliste-ID ist vorhanden, aber nicht im kanonischen Format geschrieben, das vom Standard erwartet wird.',
            'what_to_do': 'Benennen Sie die Werteliste-ID in die vom Validator gezeigte kanonische Form um.',
            'category': 'Value catalog issues',
        },
        'missing_ifc_property_uri_reference': {
            'title': 'Official IFC property URI missing',
            'what_it_means': 'Dieses Merkmal hat keine offizielle Referenz auf eine IFC property URI.',
            'what_to_do': 'Falls eine offizielle IFC property URI existiert, fügen Sie sie hinzu. Falls keine existiert, darf das Feld leer bleiben.',
            'category': 'Property definitions',
        },
        'missing_ifc_property_set_reference': {
            'title': 'Official IFC property-set reference missing',
            'what_it_means': 'Dieses Merkmal hat keine offizielle Referenz auf ein IfcPropertySet oder IfcQuantitySet.',
            'what_to_do': 'Fügen Sie die offizielle IFC-Set-Referenz hinzu, falls sie anwendbar ist; andernfalls geben Sie eine Merkmalsgruppe an.',
            'category': 'Property definitions',
        },
        'missing_property_set_locator': {
            'title': 'No property grouping provided',
            'what_it_means': 'Dieses Merkmal hat weder eine offizielle IFC property-set reference noch eine lokale Merkmalsgruppe.',
            'what_to_do': 'Geben Sie eine IfcPropertySet / IfcQuantitySet reference an oder ergänzen Sie eine Merkmalsgruppe.',
            'category': 'Property definitions',
        },
        'unknown_property_set_reference': {
            'title': 'Property set reference is unknown',
            'what_it_means': 'Das im Template verwendete PropertySet ist nicht im Merkmalskatalog registriert.',
            'what_to_do': 'Verwenden Sie einen registrierten IfcPropertySet-, IfcQuantitySet- oder Merkmalsgruppe-Wert.',
            'category': 'Data template assignment issues',
        },
        'matrix_unknown_merkmal_id': {
            'title': 'Property ID in template area is unknown',
            'what_it_means': 'Das Data Template referenziert eine Merkmal-ID, die im Property-Tab nicht registriert ist.',
            'what_to_do': 'Verwenden Sie eine Merkmal-ID, die in Merkmale existiert.',
            'category': 'Data template assignment issues',
        },
        'invalid_allowed_values_override': {
            'title': 'Selected values do not match the allowed values',
            'what_it_means': 'Die für eine Objekt-/Merkmals-Kombination ausgewählten Werte gehören nicht zu den offiziellen Allowed Values dieses Merkmals.',
            'what_to_do': 'Korrigieren Sie die Werte oder erweitern Sie den offiziellen Werte, falls die fehlenden Werte tatsächlich benötigt werden.',
            'category': 'Data template assignment issues',
        },
        'missing_source_code': {
            'title': 'Source code is missing',
            'what_it_means': 'Das Feld für source/provenance ist leer.',
            'what_to_do': 'Geben Sie einen registrierten source code an oder verwenden Sie Organisation, falls keine externe Quelle existiert.',
            'category': 'Document and source governance',
        },
        'unknown_source_code': {
            'title': 'Source code is not registered',
            'what_it_means': 'Es wurde eine source verwendet, die im document/source register nicht formell registriert ist.',
            'what_to_do': 'Registrieren Sie die source in Dokumente oder verwenden Sie Organisation.',
            'category': 'Document and source governance',
        },
        'missing_document_code': {
            'title': 'Document code is missing',
            'what_it_means': 'Die kanonische document identifier fehlt.',
            'what_to_do': 'Geben Sie den kanonischen document code gemäss Ihrer Governance-Regel an.',
            'category': 'Document and source governance',
        },
        'duplicate_document_code': {
            'title': 'Document code is used more than once',
            'what_it_means': 'Der gleiche kanonische document code kommt mehrfach vor.',
            'what_to_do': 'Stellen Sie sicher, dass jeder document code eindeutig ist.',
            'category': 'Document and source governance',
        },
        'missing_ifc_type_object_entity': {
            'title': 'IFC type mapping is missing',
            'what_it_means': 'Es existiert ein IFC object-level mapping, aber es wurde kein type-level mapping angegeben.',
            'what_to_do': 'Fügen Sie das IFC type mapping hinzu, falls es anwendbar ist; andernfalls kann es oft so belassen werden.',
            'category': 'Object definitions',
        },
        'missing_predefined_type': {
            'title': 'IFC predefined type is missing',
            'what_it_means': 'Für dieses Objekt wurde kein predefined type angegeben.',
            'what_to_do': 'Fügen Sie den predefined type hinzu, falls das IFC mapping einen solchen verlangt.',
            'category': 'Object definitions',
        },
    }
    return mapping.get(code, {
        'title': code.replace('_', ' ').capitalize(),
        'what_it_means': 'Der Validator hat ein Problem gefunden, das geprüft werden sollte.',
        'what_to_do': 'Öffnen Sie das referenzierte Sheet und die betroffene Row, prüfen Sie den Wert und korrigieren Sie ihn gemäss der Workbook-Guidance.',
        'category': 'Other validation issues',
    })


def render_markdown_report(report: dict) -> str:
    summary = report.get('summary', {})
    findings = report.get('findings', [])
    normalizations = report.get('normalizations', [])
    from collections import defaultdict
    grouped = defaultdict(list)
    for f in findings:
        meta = _layman_mapping(f['code'])
        grouped[meta['category']].append((f, meta))

    lines = []
    lines.append('# Data Dictionary Validation Report')
    lines.append('')
    lines.append(f'**Workbook:** `{report.get("workbook")}`')
    lines.append('')
    lines.append('## Gesamtergebnis')
    lines.append('')
    lines.append(f'- Blockierende Fehler: **{summary.get("errors", 0)}**')
    lines.append(f'- Warnungen: **{summary.get("warnings", 0)}**')
    lines.append(f'- Normalisierungen / abgeleitete Hinweise: **{summary.get("normalizations", 0)}**')
    lines.append(f'- Pipeline gültig: **{summary.get("pipeline_valid", False)}**')
    lines.append('')
    if summary.get('errors', 0) > 0:
        lines.append('> Das Workbook ist **noch nicht bereit für einen fehlerfreien Durchlauf**. Bitte prüfen Sie zuerst die blockierenden Fehler.')
    else:
        lines.append('> Es wurden keine blockierenden Validierungsfehler gefunden. Bitte prüfen Sie trotzdem die Warnungen und Normalisierungshinweise.')
    lines.append('')
    lines.append('## Befunde nach Thema')
    lines.append('')
    for category in sorted(grouped.keys()):
        lines.append(f'### {category}')
        lines.append('')
        for f, meta in grouped[category]:
            where = []
            if f.get('sheet'):
                where.append(f"Sheet: `{f['sheet']}`")
            if f.get('row'):
                where.append(f"Row: `{f['row']}`")
            if f.get('cell'):
                where.append(f"Cell: `{f['cell']}`")
            lines.append(f"- **{meta['title']}** ({f['level'].upper()})")
            if where:
                lines.append(f"  - Ort: {', '.join(where)}")
            lines.append(f"  - Bedeutung: {meta['what_it_means']}")
            lines.append(f"  - Empfohlene Korrektur: {meta['what_to_do']}")
            lines.append(f"  - Technisches Detail: `{f['code']}` — {f['message']}")
            lines.append('')
    lines.append('## Normalisierungshinweise')
    lines.append('')
    if not normalizations:
        lines.append('- Es wurden keine automatischen Normalisierungen oder Ableitungshinweise erfasst.')
    else:
        for n in normalizations:
            where = []
            if n.get('sheet'):
                where.append(f"Sheet: `{n['sheet']}`")
            if n.get('row'):
                where.append(f"Row: `{n['row']}`")
            if n.get('column'):
                where.append(f"Field: `{n['column']}`")
            lines.append(f"- {'; '.join(where)}")
            lines.append(f"  - Originalwert: `{n.get('original_value')}`")
            lines.append(f"  - Normalisiert / abgeleitet: `{n.get('normalized_value')}`")
            lines.append(f"  - Begründung: {n.get('reason')}")
            lines.append('')
    return '\n'.join(lines) + '\n'


def main(argv=None):
    parser = argparse.ArgumentParser(description='Validate HE-DD Strukturvorlage workbook for SchemaForge pipeline use.')
    parser.add_argument('workbook', help='Path to workbook (.xlsx)')
    parser.add_argument('--json-out', help='Optional path to write JSON report')
    parser.add_argument('--md-out', help='Optional path to write layman-readable Markdown report')
    args = parser.parse_args(argv)

    workbook = Path(args.workbook).resolve()
    validator = Validator(workbook)
    report = validator.validate()

    if args.json_out:
        out = Path(args.json_out)
        out.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    if args.md_out:
        md_out = Path(args.md_out)
        md_out.write_text(render_markdown_report(report))

    print(json.dumps(report['summary'], indent=2))
    for f in report['findings'][:50]:
        loc = []
        if f['sheet']:
            loc.append(f['sheet'])
        if f['row']:
            loc.append(f"row {f['row']}")
        where = ' @ '.join(loc)
        print(f"[{f['level'].upper()}] {f['code']}: {f['message']}" + (f" ({where})" if where else ''))
    return 1 if not report['summary']['pipeline_valid'] else 0


if __name__ == '__main__':
    raise SystemExit(main())
