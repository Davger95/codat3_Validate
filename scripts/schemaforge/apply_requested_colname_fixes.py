from openpyxl import load_workbook
from pathlib import Path

path = Path('/home/Dave/.openclaw/workspace-datadict/P_workspace_codat3_validator/templates/Strukturvorlage_DataDictionary_v2_empty.xlsx')
wb = load_workbook(path)

# Header updates
ws = wb['Header']
updates = {
    (5,3): 'SYSTEM-GENERATED',
    (5,4): 'Generated from DictionaryName (EN)',
    (6,3): 'OPTIONAL',
    (7,3): 'OPTIONAL',
    (8,3): 'REQUIRED',
    (8,4): 'Required authoring title in English — source for DictionaryCode generation',
    (10,3): 'SYSTEM-GENERATED',
    (10,4): 'Derived from OrganizationCode and DictionaryCode; canonical pattern https://example.com/<OrganizationCode>/<DictionaryCode>',
    (4,4): 'Short org code — lowercase, max 7 chars, appears in all derived IRIs',
}
for (r,c), v in updates.items():
    ws.cell(r,c).value = v

# Classes updates
ws = wb['Classes']
ws.cell(1,31).value = 'Classes.RelatedDocumentName (EN)'
ws.cell(7,11).value = 'must match Rules.Objekt-Einordnung'
ws.cell(7,27).value = 'must match Rules.Status'
ws.cell(7,31).value = 'must match a DocumentName (EN) entry from Documents.DocumentName (EN)'

# Properties updates
ws = wb['Properties']
ws.cell(1,17).value = 'Properties.EnumerationDesignation (EN)'
ws.cell(7,15).value = 'must match Rules.DataType'
ws.cell(7,16).value = 'must match Rules.IFC Data Type'
ws.cell(7,17).value = 'must match a Designation (EN) entry from Values.Designation (EN)'
ws.cell(7,28).value = 'must match Rules.Status'

# Values updates
ws = wb['Values']
ws.cell(7,16).value = 'must match Rules.Status'

# Documents updates
ws = wb['Documents']
ws.cell(7,11).value = 'must match Rules.ISG - Informationssicherheitsgesetz'
ws.cell(7,12).value = 'must match Rules.FAIR Prinzipien'
ws.cell(7,17).value = 'must match Rules.Status'

# GroupOfProperties updates
ws = wb['GroupOfProperties']
ws.cell(7,12).value = 'must match Rules.Status'

wb.save(path)
print('workbook updated')
