"""
SchemaForge — excel_reader.py
Parses all 6 tabs of an HE-SEM DD Excel file into typed dataclasses.
Handles the 4-row header structure: row1=merged title, row2=col headers,
row3=REQUIRED flags, row4=descriptions, data from row5 onward.
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import openpyxl


# ── Dataclasses ──────────────────────────────────────────────────────────────

@dataclass
class DictionaryMeta:
    """Tab 1 — key/value metadata pairs (skips section headers)."""
    org_code: str = ""
    org_name_de: str = ""
    org_name_fr: str = ""
    org_name_en: str = ""
    dd_uri: str = ""
    dd_version: str = ""
    dd_status: str = ""
    countries: str = ""
    raw: dict = field(default_factory=dict)


@dataclass
class DDClass:
    """Tab 2 — one class/object type."""
    code: str = ""
    class_type: str = "Class"
    name_de: str = ""
    name_fr: str = ""
    name_en: str = ""
    definition_de: str = ""
    definition_fr: str = ""
    owned_uri: str = ""
    parent_class_code: Optional[str] = None
    ifc_entity_code: Optional[str] = None
    ifc_predefined_type: Optional[str] = None
    ifc_uri: Optional[str] = None
    rds_reference: Optional[str] = None
    crb_code: Optional[str] = None
    status: str = "Preview"
    version_date: Optional[str] = None
    document_reference: Optional[str] = None
    synonyms: Optional[str] = None
    countries_of_use: Optional[str] = None


@dataclass
class DDProperty:
    """Tab 3 — one property."""
    code: str = ""
    name_de: str = ""
    name_fr: str = ""
    name_en: str = ""
    definition_de: str = ""
    definition_fr: str = ""
    owned_uri: str = ""
    data_type: str = "STRING"
    data_type_ifc: Optional[str] = None
    property_value_kind: str = "Single"
    unit_label: Optional[str] = None
    unit_qudt_iri: Optional[str] = None
    physical_quantity: Optional[str] = None
    min_value: Optional[str] = None
    max_value: Optional[str] = None
    enumeration_values: Optional[str] = None
    ifc_property_uri: Optional[str] = None
    ifc_pset_uri: Optional[str] = None
    property_set_name: Optional[str] = None
    rds_reference: Optional[str] = None
    status: str = "Preview"
    version_date: Optional[str] = None
    prov_attributed_to: Optional[str] = None


@dataclass
class DDClassProperty:
    """Tab 4 — assignment of a property to a class."""
    class_code: str = ""
    property_code: str = ""
    property_set_name: Optional[str] = None
    is_required: bool = False
    is_writable: bool = True
    predefined_value: Optional[str] = None
    unit_override: Optional[str] = None
    sort_number: Optional[int] = None
    loin_sia_phase: Optional[str] = None
    loin_role: Optional[str] = None
    loin_use_case: Optional[str] = None
    allowed_values_override: Optional[str] = None


@dataclass
class DDAllowedValue:
    """Tab 5 — enumeration value for a property."""
    property_code: str = ""
    code: str = ""
    value_de: str = ""
    value_fr: str = ""
    value_en: str = ""
    definition_de: Optional[str] = None
    owned_uri: str = ""
    sort_number: Optional[int] = None
    skos_exact_match: Optional[str] = None
    status: str = "Preview"


@dataclass
class DDConceptRelation:
    """Tab 6 — cross-dictionary SKOS link (property or class level)."""
    subject_code: str = ""        # code from Property or Class tab
    relation_type: str = ""       # skos:exactMatch | skos:closeMatch | etc.
    related_uri: str = ""         # full URI of related concept
    notes: Optional[str] = None


@dataclass
class DataDictionary:
    """Full in-memory representation of one HE-SEM DD Excel file."""
    source_file: Path = field(default_factory=Path)
    meta: DictionaryMeta = field(default_factory=DictionaryMeta)
    classes: list[DDClass] = field(default_factory=list)
    properties: list[DDProperty] = field(default_factory=list)
    class_properties: list[DDClassProperty] = field(default_factory=list)
    allowed_values: list[DDAllowedValue] = field(default_factory=list)
    concept_relations: list[DDConceptRelation] = field(default_factory=list)

    # Derived index: class_code → DDClass
    @property
    def class_index(self) -> dict[str, DDClass]:
        return {c.code: c for c in self.classes}

    # Derived index: property_code → DDProperty
    @property
    def property_index(self) -> dict[str, DDProperty]:
        return {p.code: p for p in self.properties}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _str(v) -> Optional[str]:
    """Coerce cell value to stripped string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _bool(v) -> bool:
    if v is None:
        return False
    return str(v).strip().upper() in ("TRUE", "YES", "1")


def _int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _read_sheet(ws, header_row: int = 2, data_start: int = 5) -> tuple[list[str], list[dict]]:
    """Return (headers, rows) where each row is a dict keyed by header."""
    headers = [cell.value for cell in ws[header_row]]
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return headers, rows


# ── Tab parsers ───────────────────────────────────────────────────────────────

def _parse_dictionary(ws) -> DictionaryMeta:
    """Tab 1: parse Field/Value pairs (skip section headers and empty rows)."""
    meta = DictionaryMeta()
    for row in ws.iter_rows(min_row=2, values_only=True):
        field_name, value = row[0], row[1]
        if field_name is None or value is None:
            continue
        fn = str(field_name).strip()
        v  = _str(value) or ""
        meta.raw[fn] = v
        # Map known fields
        fn_lower = fn.lower()
        if fn_lower == "organizationcode":
            meta.org_code = v
        elif fn_lower == "organizationname" and "(de)" in fn_lower or fn == "OrganizationName_DE":
            meta.org_name_de = v
        elif fn == "DictionaryUri" or fn_lower == "dictionaryuri":
            meta.dd_uri = v
        elif fn_lower == "version":
            meta.dd_version = v
        elif fn_lower == "status":
            meta.dd_status = v
        elif fn_lower == "countriesofuse":
            meta.countries = v
    # fallback: get org_code directly
    if not meta.org_code and "OrganizationCode" in meta.raw:
        meta.org_code = meta.raw["OrganizationCode"]
    return meta


def _parse_classes(ws) -> list[DDClass]:
    _, rows = _read_sheet(ws)
    classes = []
    for r in rows:
        c = DDClass(
            code                = _str(r.get("Code")) or "",
            class_type          = _str(r.get("ClassType")) or "Class",
            name_de             = _str(r.get("Name (DE)")) or "",
            name_fr             = _str(r.get("Name (FR)")) or "",
            name_en             = _str(r.get("Name (EN)")) or "",
            definition_de       = _str(r.get("Definition (DE)")) or "",
            definition_fr       = _str(r.get("Definition (FR)")) or "",
            owned_uri           = _str(r.get("OwnedUri")) or "",
            parent_class_code   = _str(r.get("ParentClassCode")),
            ifc_entity_code     = _str(r.get("IFC_EntityCode")),
            ifc_predefined_type = _str(r.get("IFC_PredefinedType")),
            ifc_uri             = _str(r.get("IFC_URI")),
            rds_reference       = _str(r.get("RDS_Reference")),
            crb_code            = _str(r.get("CRB_Code")),
            status              = _str(r.get("Status")) or "Preview",
            version_date        = _str(r.get("VersionDateUtc")),
            document_reference  = _str(r.get("DocumentReference")),
            synonyms            = _str(r.get("Synonyms")),
            countries_of_use    = _str(r.get("CountriesOfUse")),
        )
        if c.code:
            classes.append(c)
    return classes


def _parse_properties(ws) -> list[DDProperty]:
    _, rows = _read_sheet(ws)
    props = []
    for r in rows:
        p = DDProperty(
            code                 = _str(r.get("Code")) or "",
            name_de              = _str(r.get("Name (DE)")) or "",
            name_fr              = _str(r.get("Name (FR)")) or "",
            name_en              = _str(r.get("Name (EN)")) or "",
            definition_de        = _str(r.get("Definition (DE)")) or "",
            definition_fr        = _str(r.get("Definition (FR)")) or "",
            owned_uri            = _str(r.get("OwnedUri")) or "",
            data_type            = _str(r.get("DataType")) or "STRING",
            data_type_ifc        = _str(r.get("DataType_IFC")),
            property_value_kind  = _str(r.get("PropertyValueKind")) or "Single",
            unit_label           = _str(r.get("Unit_Label")),
            unit_qudt_iri        = _str(r.get("Unit_QUDT_IRI")),
            physical_quantity    = _str(r.get("PhysicalQuantity")),
            min_value            = _str(r.get("MinValue")),
            max_value            = _str(r.get("MaxValue")),
            enumeration_values   = _str(r.get("EnumerationValues")),
            ifc_property_uri     = _str(r.get("IFC_PropertyURI")),
            ifc_pset_uri         = _str(r.get("IFC_PsetURI")),
            property_set_name    = _str(r.get("PropertySetName")),
            rds_reference        = _str(r.get("RDS_Reference")),
            status               = _str(r.get("Status")) or "Preview",
            version_date         = _str(r.get("VersionDateUtc")),
            prov_attributed_to   = _str(r.get("PROV_AttributedTo")),
        )
        if p.code:
            props.append(p)
    return props


def _parse_class_properties(ws) -> list[DDClassProperty]:
    _, rows = _read_sheet(ws)
    cps = []
    for r in rows:
        cp = DDClassProperty(
            class_code            = _str(r.get("ClassCode")) or "",
            property_code         = _str(r.get("PropertyCode")) or "",
            property_set_name     = _str(r.get("PropertySetName")),
            is_required           = _bool(r.get("IsRequired")),
            is_writable           = _bool(r.get("IsWritable")) if r.get("IsWritable") is not None else True,
            predefined_value      = _str(r.get("PredefinedValue")),
            unit_override         = _str(r.get("Unit_Override")),
            sort_number           = _int(r.get("SortNumber")),
            loin_sia_phase        = _str(r.get("LOIN_SIA_Phase")),
            loin_role             = _str(r.get("LOIN_Role")),
            loin_use_case         = _str(r.get("LOIN_UseCase")),
            allowed_values_override = _str(r.get("AllowedValues_Override")),
        )
        if cp.class_code and cp.property_code:
            cps.append(cp)
    return cps


def _parse_allowed_values(ws) -> list[DDAllowedValue]:
    _, rows = _read_sheet(ws)
    avs = []
    for r in rows:
        av = DDAllowedValue(
            property_code  = _str(r.get("PropertyCode")) or "",
            code           = _str(r.get("Code")) or "",
            value_de       = _str(r.get("Value (DE)")) or "",
            value_fr       = _str(r.get("Value (FR)")) or "",
            value_en       = _str(r.get("Value (EN)")) or "",
            definition_de  = _str(r.get("Definition (DE)")),
            owned_uri      = _str(r.get("OwnedUri")) or "",
            sort_number    = _int(r.get("SortNumber")),
            skos_exact_match = _str(r.get("SKOS_ExactMatch")),
            status         = _str(r.get("Status")) or "Preview",
        )
        if av.property_code and av.code:
            avs.append(av)
    return avs


def _parse_concept_relations(ws) -> list[DDConceptRelation]:
    """Tab 6 — works for both 'ConceptRelation' (KBOB) and 'PropertyRelation' (bDCH) sheet names."""
    _, rows = _read_sheet(ws)
    crs = []
    for r in rows:
        cr = DDConceptRelation(
            subject_code  = _str(r.get("PropertyCode")) or "",
            relation_type = _str(r.get("RelationType")) or "",
            related_uri   = _str(r.get("RelatedPropertyUri")) or "",
            notes         = _str(r.get("Notes")),
        )
        if cr.subject_code and cr.relation_type and cr.related_uri:
            crs.append(cr)
    return crs


# ── Public API ────────────────────────────────────────────────────────────────

def load_dd(path: Path) -> DataDictionary:
    """Load a complete HE-SEM DD Excel file into a DataDictionary object."""
    path = Path(path)
    wb = openpyxl.load_workbook(path)

    # Tab 6 may be named ConceptRelation or PropertyRelation
    tab6_name = None
    for candidate in ("ConceptRelation", "PropertyRelation"):
        if candidate in wb.sheetnames:
            tab6_name = candidate
            break

    dd = DataDictionary(
        source_file      = path,
        meta             = _parse_dictionary(wb["Dictionary"]),
        classes          = _parse_classes(wb["Class"]),
        properties       = _parse_properties(wb["Property"]),
        class_properties = _parse_class_properties(wb["ClassProperty"]),
        allowed_values   = _parse_allowed_values(wb["AllowedValue"]),
        concept_relations = _parse_concept_relations(wb[tab6_name]) if tab6_name else [],
    )
    return dd


if __name__ == "__main__":
    import sys
    from pathlib import Path
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if target:
        dd = load_dd(target)
        print(f"Loaded: {target.name}")
        print(f"  org_code:         {dd.meta.org_code}")
        print(f"  classes:          {len(dd.classes)}")
        print(f"  properties:       {len(dd.properties)}")
        print(f"  class_properties: {len(dd.class_properties)}")
        print(f"  allowed_values:   {len(dd.allowed_values)}")
        print(f"  concept_relations:{len(dd.concept_relations)}")
        print("\n  Classes:")
        for c in dd.classes:
            print(f"    {c.code:25} IFC={c.ifc_entity_code} PredType={c.ifc_predefined_type}")
        print("\n  ConceptRelations:")
        for cr in dd.concept_relations:
            print(f"    {cr.subject_code:15} {cr.relation_type:20} {cr.related_uri}")
